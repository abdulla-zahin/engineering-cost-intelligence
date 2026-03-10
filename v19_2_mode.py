
from datetime import date
from io import BytesIO
from pathlib import Path
from openpyxl.utils import get_column_letter
import textwrap
import sqlite3

import matplotlib
import matplotlib.pyplot as plt
matplotlib.use("Agg")
import pandas as pd
import streamlit as st
from matplotlib.backends.backend_pdf import PdfPages


st.set_page_config(
    page_title="Project Cost Intelligence System",
    page_icon="🏗️",
    layout="wide",
)

PROJECT_TYPE_SPLITS = {
    "General Contracting": {
        "Materials": 40,
        "Labor": 25,
        "Transportation": 5,
        "Office Expense": 5,
        "Salaries / Overheads": 10,
        "Company Profit": 10,
        "Contingency": 5,
    },
    "Pipeline Project": {
        "Materials": 50,
        "Labor": 20,
        "Transportation": 10,
        "Office Expense": 5,
        "Salaries / Overheads": 5,
        "Company Profit": 7,
        "Contingency": 3,
    },
    "Civil Construction": {
        "Materials": 45,
        "Labor": 30,
        "Transportation": 5,
        "Office Expense": 5,
        "Salaries / Overheads": 8,
        "Company Profit": 5,
        "Contingency": 2,
    },
    "Mechanical Installation": {
        "Materials": 38,
        "Labor": 32,
        "Transportation": 6,
        "Office Expense": 5,
        "Salaries / Overheads": 9,
        "Company Profit": 7,
        "Contingency": 3,
    },
    "Maintenance Contract": {
        "Materials": 25,
        "Labor": 40,
        "Transportation": 8,
        "Office Expense": 7,
        "Salaries / Overheads": 10,
        "Company Profit": 7,
        "Contingency": 3,
    },
    "EPC Project": {
        "Materials": 42,
        "Labor": 22,
        "Transportation": 6,
        "Office Expense": 6,
        "Salaries / Overheads": 12,
        "Company Profit": 8,
        "Contingency": 4,
    },
}

CURRENCY_SYMBOLS = {
    "AED": "AED",
    "USD": "$",
    "INR": "₹",
}


def reset_allocation_widget_state(project_type: str) -> None:
    """Reset allocation defaults and visible widget values for the selected project type."""
    st.session_state.allocations_state = PROJECT_TYPE_SPLITS[project_type].copy()
    for category, value in st.session_state.allocations_state.items():
        st.session_state[f"alloc_{category}"] = float(value)
    st.session_state.last_project_type = project_type


REQUIRED_BOQ_COLUMNS = ["Item", "Quantity", "Unit Cost", "Category"]
BOQ_COLUMN_ALIASES = {
    "Item": ["item", "description", "item description", "boq item", "work item", "scope item"],
    "Quantity": ["quantity", "qty", "qty.", "qnty", "quantities"],
    "Unit Cost": ["unit cost", "unit price", "rate", "price", "unit rate", "cost per unit"],
    "Category": ["category", "type", "cost category", "group", "trade", "classification"],
}


def load_boq_file(uploaded_file) -> pd.DataFrame | None:
    if uploaded_file is None:
        return None

    file_name = uploaded_file.name.lower()
    if file_name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    if file_name.endswith(".xlsx"):
        return pd.read_excel(uploaded_file)
    raise ValueError("Unsupported file type. Please upload CSV or XLSX.")


def infer_boq_columns(raw_df: pd.DataFrame) -> dict[str, str]:
    normalized_columns = {str(col).strip().lower(): str(col) for col in raw_df.columns}
    inferred = {}

    for canonical_name, aliases in BOQ_COLUMN_ALIASES.items():
        exact_match = normalized_columns.get(canonical_name.lower())
        if exact_match:
            inferred[canonical_name] = exact_match
            continue

        for alias in aliases:
            matched = normalized_columns.get(alias.lower())
            if matched:
                inferred[canonical_name] = matched
                break

    missing = [col for col in REQUIRED_BOQ_COLUMNS if col not in inferred]
    if missing:
        available = ", ".join(map(str, raw_df.columns))
        raise ValueError(
            f"Could not identify required BOQ columns: {', '.join(missing)}. "
            f"Available columns found: {available}"
        )

    return inferred


def normalize_boq_category(value: str) -> str:
    text = str(value).strip().lower()
    mapping = {
        "material": "Materials",
        "materials": "Materials",
        "mat": "Materials",
        "labour": "Labor",
        "labor": "Labor",
        "manpower": "Labor",
        "transport": "Transportation",
        "transportation": "Transportation",
        "logistics": "Transportation",
        "delivery": "Transportation",
        "office": "Office Expense",
        "office expense": "Office Expense",
        "office expenses": "Office Expense",
        "admin": "Office Expense",
        "administration": "Office Expense",
        "salary": "Salaries / Overheads",
        "salaries": "Salaries / Overheads",
        "overhead": "Salaries / Overheads",
        "overheads": "Salaries / Overheads",
        "salaries / overheads": "Salaries / Overheads",
        "salary / overheads": "Salaries / Overheads",
        "profit": "Company Profit",
        "company profit": "Company Profit",
        "contingency": "Contingency",
        "misc": "Contingency",
        "other": "Contingency",
    }
    return mapping.get(text, str(value).strip().title())


def clean_boq_dataframe(raw_df: pd.DataFrame) -> pd.DataFrame:
    column_mapping = infer_boq_columns(raw_df)
    boq_df = raw_df.rename(columns={source: target for target, source in column_mapping.items()}).copy().dropna(how="all")

    boq_df["Item"] = boq_df["Item"].astype(str).str.strip()
    boq_df["Category"] = boq_df["Category"].astype(str).str.strip()
    boq_df["Quantity"] = pd.to_numeric(boq_df["Quantity"], errors="coerce")
    boq_df["Unit Cost"] = pd.to_numeric(boq_df["Unit Cost"], errors="coerce")
    boq_df = boq_df.dropna(subset=["Quantity", "Unit Cost"])
    boq_df = boq_df[boq_df["Item"] != ""]
    boq_df["Normalized Category"] = boq_df["Category"].map(normalize_boq_category)
    boq_df["Total Cost"] = boq_df["Quantity"] * boq_df["Unit Cost"]
    return boq_df.reset_index(drop=True)


def build_boq_comparison_dataframe(total_budget: float, allocations: dict, boq_df: pd.DataFrame) -> pd.DataFrame:
    planned_rows = []
    actual_summary = boq_df.groupby("Normalized Category")["Total Cost"].sum() if not boq_df.empty else pd.Series(dtype=float)

    for category, pct in allocations.items():
        planned_amount = total_budget * float(pct) / 100
        actual_amount = float(actual_summary.get(category, 0.0))
        planned_rows.append(
            {
                "Category": category,
                "Planned Budget": round(planned_amount, 2),
                "Actual BOQ Cost": round(actual_amount, 2),
                "Difference": round(planned_amount - actual_amount, 2),
            }
        )

    return pd.DataFrame(planned_rows)


def analyze_boq_items(boq_df: pd.DataFrame) -> tuple[dict, pd.Series]:
    category_summary = boq_df.groupby("Normalized Category")["Total Cost"].sum().sort_values(ascending=False)
    total_cost = float(boq_df["Total Cost"].sum())
    if boq_df.empty:
        summary = {
            "total_cost": 0.0,
            "total_items": 0,
            "most_expensive_item": "N/A",
            "most_expensive_cost": 0.0,
        }
    else:
        idx = boq_df["Total Cost"].idxmax()
        summary = {
            "total_cost": total_cost,
            "total_items": int(len(boq_df)),
            "most_expensive_item": str(boq_df.loc[idx, "Item"]),
            "most_expensive_cost": float(boq_df.loc[idx, "Total Cost"]),
        }
    return summary, category_summary


def build_separate_case_summary(case_results: list[dict]) -> pd.DataFrame:
    if not case_results:
        return pd.DataFrame()

    rows = []
    for case in case_results:
        summary = case["summary"]
        rows.append(
            {
                "BOQ File": case["file_name"],
                "Total Cost": float(summary.get("total_cost", 0.0)),
                "BOQ Items": int(summary.get("total_items", 0)),
                "Highest Item": str(summary.get("most_expensive_item", "N/A")),
                "Highest Item Cost": float(summary.get("most_expensive_cost", 0.0)),
            }
        )

    return pd.DataFrame(rows)


def get_allocation_status(total_pct: float) -> str:
    if abs(total_pct - 100) < 1e-9:
        return "Balanced"
    if total_pct < 100:
        return "Under Allocated"
    return "Over Allocated"


def build_budget_dataframe(total_budget: float, allocations: dict) -> pd.DataFrame:
    rows = []
    for category, pct in allocations.items():
        rows.append(
            {
                "Category": category,
                "Allocation %": round(float(pct), 2),
                "Amount": round(total_budget * float(pct) / 100, 2),
            }
        )
    return pd.DataFrame(rows)


def analyze_budget(allocations: dict, total_pct: float) -> list[str]:
    insights = []
    status = get_allocation_status(total_pct)

    if status == "Balanced":
        insights.append("✅ Total allocation is perfectly balanced at 100%.")
    elif status == "Under Allocated":
        insights.append("⚠️ Total allocation is below 100%. Some budget is still unassigned.")
    else:
        insights.append("❌ Total allocation exceeds 100%. Adjust the percentages to match the budget.")

    profit = allocations["Company Profit"]
    labor = allocations["Labor"]
    materials = allocations["Materials"]
    contingency = allocations["Contingency"]
    transport = allocations["Transportation"]

    if profit < 8:
        insights.append("⚠️ Profit margin is below 8%. This may be too tight for comfortable execution.")
    elif profit <= 15:
        insights.append("✅ Profit margin looks healthy for a practical contracting estimate.")
    else:
        insights.append("⚠️ Profit margin is quite high. Double-check competitiveness and client expectations.")

    if labor < 20:
        insights.append("⚠️ Labor allocation is low. Execution quality or manpower coverage may suffer.")
    elif labor > 35:
        insights.append("⚠️ Labor allocation is very high. Verify productivity assumptions and manpower planning.")

    if materials < 30:
        insights.append("⚠️ Material allocation is low. Review BOQ realism and procurement assumptions.")
    elif materials > 60:
        insights.append("⚠️ Material allocation is very high. Check whether supply cost is dominating the project.")

    if contingency < 5:
        insights.append("⚠️ Contingency is low. This leaves little room for site surprises or change orders.")
    else:
        insights.append("✅ Contingency reserve provides a reasonable safety buffer.")

    if transport > 10:
        insights.append("⚠️ Transportation cost is high. Recheck logistics, fuel, and delivery assumptions.")

    return insights


def generate_report_text(
    project_name: str,
    project_type: str,
    company_name: str,
    client_name: str,
    project_reference: str,
    currency_symbol: str,
    total_budget: float,
    allocations: dict,
    total_pct: float,
) -> dict[str, str]:
    profit_pct = allocations["Company Profit"]
    profit_amount = total_budget * profit_pct / 100
    allocation_status = get_allocation_status(total_pct).lower()

    estimate_summary = (
        f"Project Reference: {project_reference}. Client: {client_name}. "
        f"Based on the submitted project details for {project_name}, categorized under {project_type}, "
        f"{company_name} has prepared a preliminary budget estimate with a total projected value of "
        f"{currency_symbol} {total_budget:,.2f}. The estimate has been distributed across key execution heads "
        f"including materials, labor, transportation, office expenses, overheads, contingency, and company profit. "
        f"The current model reflects an expected profit margin of {profit_pct:.2f}% with an estimated profit value "
        f"of {currency_symbol} {profit_amount:,.2f}. Overall allocation status is presently {allocation_status}, "
        f"subject to final technical and commercial review."
    )

    proposal_note = (
        f"Reference {project_reference} has been prepared for {client_name}. "
        f"We are pleased to submit this preliminary commercial budget consideration for {project_name}. "
        f"This estimate has been developed based on the selected project type, expected execution requirements, "
        f"material demand, labor deployment, logistics considerations, and administrative overheads. "
        f"The proposed allocation is intended to maintain practical project execution while preserving commercial "
        f"viability for {company_name}. Any final quotation or commercial submission should be validated against "
        f"approved scope, drawings, specifications, and prevailing market rates before issue."
    )

    disclaimer_note = (
        "This budget estimate is prepared solely for planning, review, and preliminary commercial evaluation purposes. "
        "Final values may vary depending on site conditions, actual quantities, client requirements, specification "
        "changes, transportation conditions, authority approvals, and material price fluctuations at the time of execution. "
        "Any variation in scope, timeline, procurement conditions, or execution methodology may lead to revision "
        "of the final commercial offer and agreement terms."
    )

    internal_note = (
        f"Internal review reference: {project_reference}. "
        f"From an internal review standpoint, the current allocation for {project_name} appears "
        f"{allocation_status} with a projected profit margin of {profit_pct:.2f}%. "
        "Management is advised to review category distribution, execution risks, and contingency adequacy "
        "before final approval or client submission."
    )

    return {
        "Estimate Summary": estimate_summary,
        "Commercial Proposal Note": proposal_note,
        "Terms / Disclaimer Note": disclaimer_note,
        "Internal Approval Note": internal_note,
    }


def create_pdf_report(
    project_name: str,
    project_type: str,
    company_name: str,
    client_name: str,
    project_reference: str,
    currency_code: str,
    currency_symbol: str,
    author_name: str,
    today: str,
    total_budget: float,
    allocations: dict,
    df: pd.DataFrame,
    insights: list[str],
    report_sections: dict[str, str],
    boq_df: pd.DataFrame | None = None,
    boq_compare_df: pd.DataFrame | None = None,
    boq_summary: dict | None = None,
    project_history_df: pd.DataFrame | None = None,
    allocation_history_df: pd.DataFrame | None = None,
    comparison_history_df: pd.DataFrame | None = None,
) -> BytesIO:
    pdf_buffer = BytesIO()
    profit_amount = total_budget * allocations["Company Profit"] / 100
    total_pct = sum(allocations.values())
    allocation_status = get_allocation_status(total_pct)
    boq_df = boq_df if boq_df is not None else pd.DataFrame()
    boq_compare_df = boq_compare_df if boq_compare_df is not None else pd.DataFrame()
    boq_summary = boq_summary or {"total_cost": 0.0, "total_items": 0, "most_expensive_item": "N/A", "most_expensive_cost": 0.0}
    project_history_df = project_history_df if project_history_df is not None else pd.DataFrame()
    allocation_history_df = allocation_history_df if allocation_history_df is not None else pd.DataFrame()
    comparison_history_df = comparison_history_df if comparison_history_df is not None else pd.DataFrame()

    with PdfPages(pdf_buffer) as pdf:
        fig1 = plt.figure(figsize=(8.27, 11.69))
        ax1 = fig1.add_axes([0, 0, 1, 1])
        ax1.axis("off")

        y = 0.96
        ax1.text(0.5, y, company_name, ha="center", va="top", fontsize=20, fontweight="bold")
        y -= 0.04
        ax1.text(0.5, y, "PROJECT COST INTELLIGENCE REPORT", ha="center", va="top", fontsize=14, fontweight="bold")
        y -= 0.06

        left_lines = [
            f"Project Name: {project_name}",
            f"Client Name: {client_name}",
            f"Project Reference: {project_reference}",
            f"Project Type: {project_type}",
            f"Date: {today}",
            f"Prepared By: {author_name}",
            f"Currency: {currency_code}",
        ]
        right_lines = [
            f"Total Budget: {currency_symbol} {total_budget:,.2f}",
            f"Profit Margin: {allocations['Company Profit']:.2f}%",
            f"Profit Amount: {currency_symbol} {profit_amount:,.2f}",
            f"Allocation Status: {allocation_status}",
        ]

        y_left = y
        for line in left_lines:
            ax1.text(0.08, y_left, line, fontsize=10, va="top")
            y_left -= 0.03

        y_right = y
        for line in right_lines:
            ax1.text(0.58, y_right, line, fontsize=10, va="top", fontweight="bold")
            y_right -= 0.04

        y = min(y_left, y_right) - 0.03
        ax1.text(0.08, y, "Budget Intelligence", fontsize=12, fontweight="bold", va="top")
        y -= 0.03

        for insight in insights:
            wrapped = textwrap.fill(insight, width=90)
            ax1.text(0.09, y, wrapped, fontsize=9, va="top")
            y -= 0.04 + 0.012 * wrapped.count("\n")

        pdf.savefig(fig1, bbox_inches="tight")
        plt.close(fig1)

        fig2 = plt.figure(figsize=(8.27, 11.69))
        ax2 = fig2.add_axes([0.05, 0.05, 0.9, 0.9])
        ax2.axis("off")
        ax2.set_title("Detailed Budget Breakdown", fontsize=14, fontweight="bold", pad=20)

        display_df = df.copy()
        display_df["Allocation %"] = display_df["Allocation %"].map(lambda x: f"{x:.2f}%")
        display_df["Amount"] = display_df["Amount"].map(lambda x: f"{currency_symbol} {x:,.2f}")

        table = ax2.table(
            cellText=display_df.values,
            colLabels=display_df.columns,
            cellLoc="center",
            colLoc="center",
            loc="upper center",
        )
        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1, 1.6)

        for (row, _col), cell in table.get_celld().items():
            if row == 0:
                cell.set_text_props(weight="bold")
                cell.set_facecolor("#D9EAF7")
            elif row > 0:
                category = display_df.iloc[row - 1, 0]
                if category == "Company Profit":
                    cell.set_facecolor("#FFF2CC")
                elif category == "Contingency":
                    cell.set_facecolor("#E2F0D9")

        pdf.savefig(fig2, bbox_inches="tight")
        plt.close(fig2)

        fig3 = plt.figure(figsize=(8.27, 11.69))
        ax3 = fig3.add_axes([0.08, 0.53, 0.84, 0.36])
        ax3.pie(df["Amount"], labels=df["Category"], autopct="%1.1f%%", startangle=90)
        ax3.set_title("Budget Distribution")

        ax4 = fig3.add_axes([0.08, 0.09, 0.84, 0.32])
        ax4.bar(df["Category"], df["Amount"])
        ax4.set_title(f"Category-wise Budget Amount ({currency_code})")
        ax4.set_ylabel(currency_code)
        ax4.tick_params(axis="x", rotation=45)

        pdf.savefig(fig3, bbox_inches="tight")
        plt.close(fig3)

        fig4 = plt.figure(figsize=(8.27, 11.69))
        ax5 = fig4.add_axes([0, 0, 1, 1])
        ax5.axis("off")
        ax5.text(0.5, 0.96, "AUTO-GENERATED REPORT DRAFT", ha="center", va="top", fontsize=14, fontweight="bold")

        y = 0.9
        for title, content in report_sections.items():
            ax5.text(0.07, y, title, fontsize=11, fontweight="bold", va="top")
            y -= 0.03
            wrapped = textwrap.fill(content, width=100)
            ax5.text(0.08, y, wrapped, fontsize=9, va="top")
            y -= 0.08 + 0.014 * wrapped.count("\n")

        pdf.savefig(fig4, bbox_inches="tight")
        plt.close(fig4)

        if not boq_df.empty:
            fig5 = plt.figure(figsize=(8.27, 11.69))
            ax6 = fig5.add_axes([0, 0, 1, 1])
            ax6.axis("off")
            ax6.text(0.5, 0.96, "UPLOADED BOQ SUMMARY", ha="center", va="top", fontsize=14, fontweight="bold")

            boq_lines = [
                f"BOQ Total Cost: {currency_symbol} {boq_summary['total_cost']:,.2f}",
                f"BOQ Item Count: {boq_summary['total_items']}",
                f"Highest BOQ Item: {boq_summary['most_expensive_item']}",
                f"Highest BOQ Item Cost: {currency_symbol} {boq_summary['most_expensive_cost']:,.2f}",
            ]
            y = 0.88
            for line in boq_lines:
                ax6.text(0.08, y, line, fontsize=11, va="top")
                y -= 0.05

            boq_display_columns = ["BOQ Source", "Item", "Quantity", "Unit Cost", "Normalized Category", "Total Cost"] if "BOQ Source" in boq_df.columns else ["Item", "Quantity", "Unit Cost", "Normalized Category", "Total Cost"]
            boq_display = boq_df[boq_display_columns].copy().head(18)
            boq_display["Quantity"] = boq_display["Quantity"].map(lambda x: f"{x:,.2f}")
            boq_display["Unit Cost"] = boq_display["Unit Cost"].map(lambda x: f"{currency_symbol} {x:,.2f}")
            boq_display["Total Cost"] = boq_display["Total Cost"].map(lambda x: f"{currency_symbol} {x:,.2f}")
            boq_display.columns = ["BOQ Source", "Item", "Qty", "Unit Cost", "Category", "Total Cost"] if "BOQ Source" in boq_df.columns else ["Item", "Qty", "Unit Cost", "Category", "Total Cost"]

            table_ax = fig5.add_axes([0.06, 0.08, 0.88, 0.52])
            table_ax.axis("off")
            boq_table = table_ax.table(
                cellText=boq_display.values,
                colLabels=boq_display.columns,
                cellLoc="center",
                colLoc="center",
                loc="upper center",
            )
            boq_table.auto_set_font_size(False)
            boq_table.set_fontsize(8)
            boq_table.scale(1, 1.4)

            for (row, _col), cell in boq_table.get_celld().items():
                if row == 0:
                    cell.set_text_props(weight="bold")
                    cell.set_facecolor("#D9EAF7")

            pdf.savefig(fig5, bbox_inches="tight")
            plt.close(fig5)

        if not boq_compare_df.empty:
            fig6 = plt.figure(figsize=(8.27, 11.69))
            ax7 = fig6.add_axes([0.08, 0.58, 0.84, 0.28])
            x_positions = range(len(boq_compare_df))
            width = 0.38
            ax7.bar([x - width / 2 for x in x_positions], boq_compare_df["Planned Budget"], width=width, label="Planned")
            ax7.bar([x + width / 2 for x in x_positions], boq_compare_df["Actual BOQ Cost"], width=width, label="Actual BOQ")
            ax7.set_title("Planned Budget vs Actual BOQ")
            ax7.set_ylabel(f"Amount ({currency_code})")
            ax7.set_xticks(list(x_positions))
            ax7.set_xticklabels(boq_compare_df["Category"], rotation=45, ha="right")
            ax7.legend()

            ax8 = fig6.add_axes([0.06, 0.08, 0.88, 0.36])
            ax8.axis("off")
            compare_display = boq_compare_df.copy()
            for col in ["Planned Budget", "Actual BOQ Cost", "Difference"]:
                compare_display[col] = compare_display[col].map(lambda x: f"{currency_symbol} {x:,.2f}")

            compare_table = ax8.table(
                cellText=compare_display.values,
                colLabels=compare_display.columns,
                cellLoc="center",
                colLoc="center",
                loc="upper center",
            )
            compare_table.auto_set_font_size(False)
            compare_table.set_fontsize(8)
            compare_table.scale(1, 1.5)

            for (row, _col), cell in compare_table.get_celld().items():
                if row == 0:
                    cell.set_text_props(weight="bold")
                    cell.set_facecolor("#D9EAF7")

            pdf.savefig(fig6, bbox_inches="tight")
            plt.close(fig6)

    pdf_buffer.seek(0)
    return pdf_buffer


def create_excel_report(
    project_name: str,
    project_type: str,
    company_name: str,
    client_name: str,
    project_reference: str,
    currency_code: str,
    currency_symbol: str,
    author_name: str,
    today: str,
    total_budget: float,
    allocations: dict,
    df: pd.DataFrame,
    insights: list[str],
    report_sections: dict[str, str],
    boq_df: pd.DataFrame | None = None,
    boq_compare_df: pd.DataFrame | None = None,
    boq_summary: dict | None = None,
    logo_path: str | None = None,
    project_history_df: pd.DataFrame | None = None,
    allocation_history_df: pd.DataFrame | None = None,
    comparison_history_df: pd.DataFrame | None = None,
) -> BytesIO:
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    output = BytesIO()
    profit_pct = allocations["Company Profit"]
    profit_amount = total_budget * profit_pct / 100
    total_allocation_pct = round(sum(allocations.values()), 2)
    allocation_status = get_allocation_status(total_allocation_pct)
    insights_with_status = [f"Allocation Status: {allocation_status}"] + insights
    boq_df = boq_df if boq_df is not None else pd.DataFrame()
    boq_compare_df = boq_compare_df if boq_compare_df is not None else pd.DataFrame()
    boq_summary = boq_summary or {"total_cost": 0.0, "total_items": 0, "most_expensive_item": "N/A", "most_expensive_cost": 0.0}

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Budget Breakdown", startrow=7)
        pd.DataFrame({"Budget Insights": insights_with_status}).to_excel(
            writer, index=False, sheet_name="Budget Insights", startrow=3
        )
        if not boq_df.empty:
            boq_export_columns = ["BOQ Source", "Item", "Quantity", "Unit Cost", "Category", "Normalized Category", "Total Cost"] if "BOQ Source" in boq_df.columns else ["Item", "Quantity", "Unit Cost", "Category", "Normalized Category", "Total Cost"]
            boq_export_df = boq_df[boq_export_columns].copy()
            boq_export_df.to_excel(writer, index=False, sheet_name="BOQ Details", startrow=7)
        if not boq_compare_df.empty:
            boq_compare_df.to_excel(writer, index=False, sheet_name="BOQ Comparison", startrow=6)
        wb = writer.book

        dark_fill = PatternFill("solid", fgColor="1F4E78")
        mid_fill = PatternFill("solid", fgColor="D9EAF7")
        soft_fill = PatternFill("solid", fgColor="EDF4FB")
        gold_fill = PatternFill("solid", fgColor="FFF2CC")
        green_fill = PatternFill("solid", fgColor="E2F0D9")
        red_fill = PatternFill("solid", fgColor="FDE9E7")
        gray_fill = PatternFill("solid", fgColor="F3F3F3")

        title_font = Font(color="FFFFFF", bold=True, size=16)
        section_font = Font(bold=True, size=11)
        header_font = Font(bold=True)
        big_font = Font(bold=True, size=13)

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

        thin = Side(style="thin", color="B7B7B7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws = wb.create_sheet("Executive Summary", 0)
        ws.sheet_view.showGridLines = False
        for col, width in {"A": 16, "B": 24, "C": 24, "D": 24, "E": 24, "F": 24}.items():
            ws.column_dimensions[col].width = width

        ws["A1"].fill = dark_fill
        ws["A1"].border = border
        ws["A1"].alignment = center
        ws.merge_cells("B1:F1")
        ws["B1"] = f"{company_name} - PROJECT COST INTELLIGENCE REPORT"
        ws["B1"].fill = dark_fill
        ws["B1"].font = title_font
        ws["B1"].alignment = center
        ws.row_dimensions[1].height = 34

        if logo_path and Path(logo_path).exists():
            try:
                logo = XLImage(logo_path)
                logo.width = 72
                logo.height = 72
                ws.add_image(logo, "A1")
            except Exception:
                ws["A1"] = "LOGO"
                ws["A1"].font = header_font
                ws["A1"].alignment = center
        else:
            ws["A1"] = "LOGO"
            ws["A1"].font = header_font
            ws["A1"].alignment = center

        ws.merge_cells("A3:C3")
        ws["A3"] = "PROJECT INFORMATION"
        ws["A3"].fill = mid_fill
        ws["A3"].font = section_font
        ws["A3"].alignment = center

        ws.merge_cells("D3:F3")
        ws["D3"] = "KEY FINANCIAL METRICS"
        ws["D3"].fill = gold_fill
        ws["D3"].font = section_font
        ws["D3"].alignment = center

        info_rows = [
            ("Project Name", project_name),
            ("Project Reference", project_reference),
            ("Client Name", client_name),
            ("Project Type", project_type),
            ("Company Name", company_name),
            ("Author", author_name),
            ("Date", today),
            ("Currency", currency_code),
        ]
        metric_rows = [
            ("Total Budget", total_budget),
            ("Profit %", profit_pct / 100),
            ("Profit Amount", profit_amount),
            ("Allocation %", total_allocation_pct / 100),
            ("Status", allocation_status),
        ]

        start_row = 5
        for i, (label, value) in enumerate(info_rows, start=start_row):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"A{i}"].fill = soft_fill
            ws[f"A{i}"].font = header_font
            ws[f"A{i}"].border = border
            ws[f"B{i}"].border = border
            ws[f"C{i}"].border = border
            ws[f"A{i}"].alignment = left
            ws[f"B{i}"].alignment = left

        for i, (label, value) in enumerate(metric_rows, start=start_row):
            ws[f"D{i}"] = label
            ws[f"E{i}"] = value
            ws[f"F{i}"] = ""
            ws[f"D{i}"].fill = gray_fill
            ws[f"D{i}"].font = header_font
            ws[f"D{i}"].border = border
            ws[f"E{i}"].border = border
            ws[f"F{i}"].border = border
            ws[f"D{i}"].alignment = left
            ws[f"E{i}"].alignment = center if label != "Status" else left

            if label in {"Total Budget", "Profit Amount"}:
                ws[f"E{i}"].number_format = f'"{currency_symbol}" #,##0.00'
                ws[f"E{i}"].font = big_font
            elif label in {"Profit %", "Allocation %"}:
                ws[f"E{i}"].number_format = "0.00%"
                ws[f"E{i}"].font = big_font
            elif label == "Status":
                ws[f"E{i}"].fill = green_fill if allocation_status == "Balanced" else red_fill
                ws[f"E{i}"].font = header_font

        note_header_row = start_row + max(len(info_rows), len(metric_rows)) + 2
        note_body_row = note_header_row + 1
        ws.merge_cells(f"A{note_header_row}:F{note_header_row}")
        ws[f"A{note_header_row}"] = "EXECUTIVE NOTE"
        ws[f"A{note_header_row}"].fill = mid_fill
        ws[f"A{note_header_row}"].font = section_font
        ws[f"A{note_header_row}"].alignment = center

        note_text = (
            f"This report summarizes the recommended allocation for a {project_type.lower()} under {company_name} "
            f"for client {client_name}. Current allocation status is {allocation_status.lower()}, with an estimated "
            f"profit of {profit_pct:.2f}% and profit amount of {currency_symbol} {profit_amount:,.2f}."
        )
        ws.merge_cells(f"A{note_body_row}:F{note_body_row + 2}")
        ws[f"A{note_body_row}"] = note_text
        ws[f"A{note_body_row}"].alignment = wrap_left
        ws[f"A{note_body_row}"].border = border

        bd = wb["Budget Breakdown"]
        bd.sheet_view.showGridLines = False
        for col, width in {"A": 30, "B": 16, "C": 22, "D": 18}.items():
            bd.column_dimensions[col].width = width

        bd.merge_cells("A1:D1")
        bd["A1"] = "DETAILED BUDGET BREAKDOWN"
        bd["A1"].fill = dark_fill
        bd["A1"].font = title_font
        bd["A1"].alignment = center

        bd["A3"] = "Project Name"
        bd["B3"] = project_name
        bd["A4"] = "Project Reference"
        bd["B4"] = project_reference
        bd["A5"] = "Client Name"
        bd["B5"] = client_name
        bd["A6"] = "Project Type"
        bd["B6"] = project_type
        bd["D3"] = "Prepared By"
        bd["D4"] = author_name
        bd["D5"] = "Prepared On"
        bd["D6"] = today

        for cell in ["A3", "A4", "A5", "A6", "D3", "D5"]:
            bd[cell].fill = soft_fill
            bd[cell].font = header_font
        for cell in ["A3", "B3", "A4", "B4", "A5", "B5", "A6", "B6", "D3", "D4", "D5", "D6"]:
            bd[cell].border = border

        table_header_row = 8
        bd["D8"] = "Remarks"
        for cell in bd[table_header_row]:
            cell.fill = mid_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for row in range(table_header_row + 1, table_header_row + 1 + len(df)):
            bd[f"B{row}"].number_format = "0.00"
            bd[f"C{row}"].number_format = f'"{currency_symbol}" #,##0.00'
            bd[f"D{row}"] = "Reviewed"
            for col in "ABCD":
                bd[f"{col}{row}"].border = border

            category = bd[f"A{row}"].value
            if category == "Company Profit":
                for col in "ABCD":
                    bd[f"{col}{row}"].fill = gold_fill
                    bd[f"{col}{row}"].font = header_font
            elif category == "Contingency":
                for col in "ABCD":
                    bd[f"{col}{row}"].fill = green_fill

        total_row = table_header_row + 1 + len(df)
        bd[f"A{total_row}"] = "Total Allocation"
        bd[f"B{total_row}"] = total_allocation_pct
        bd[f"C{total_row}"] = df["Amount"].sum()
        bd[f"D{total_row}"] = allocation_status
        for col in "ABCD":
            bd[f"{col}{total_row}"].fill = gray_fill
            bd[f"{col}{total_row}"].font = header_font
            bd[f"{col}{total_row}"].border = border

        bd[f"B{total_row}"].number_format = "0.00"
        bd[f"C{total_row}"].number_format = f'"{currency_symbol}" #,##0.00'
        bd.freeze_panes = "A8"

        ins = wb["Budget Insights"]
        ins.sheet_view.showGridLines = False
        ins.column_dimensions["A"].width = 110
        ins["A1"] = "BUDGET INTELLIGENCE & RISK NOTES"
        ins["A1"].fill = dark_fill
        ins["A1"].font = title_font
        ins["A1"].alignment = center
        ins["A4"].fill = mid_fill
        ins["A4"].font = header_font
        ins["A4"].border = border

        for row in range(5, 5 + len(insights_with_status)):
            ins[f"A{row}"].border = border
            ins[f"A{row}"].alignment = wrap_left
            value = str(ins[f"A{row}"].value)
            if "✅" in value or "Balanced" in value:
                ins[f"A{row}"].fill = green_fill
            elif "⚠️" in value or "❌" in value or "Over" in value or "Under" in value:
                ins[f"A{row}"].fill = red_fill

        ins.freeze_panes = "A4"

        rpt = wb.create_sheet("Report Draft")
        rpt.sheet_view.showGridLines = False
        rpt.column_dimensions["A"].width = 28
        rpt.column_dimensions["B"].width = 110
        rpt.merge_cells("A1:B1")
        rpt["A1"] = "AUTO-GENERATED REPORT DRAFT"
        rpt["A1"].fill = dark_fill
        rpt["A1"].font = title_font
        rpt["A1"].alignment = center

        report_row = 3
        for title, content in report_sections.items():
            rpt[f"A{report_row}"] = title
            rpt[f"A{report_row}"].fill = mid_fill
            rpt[f"A{report_row}"].font = header_font
            rpt[f"A{report_row}"].border = border
            rpt[f"B{report_row}"] = content
            rpt[f"B{report_row}"].alignment = wrap_left
            rpt[f"B{report_row}"].border = border
            rpt.row_dimensions[report_row].height = 54
            report_row += 2

        ch = wb.create_sheet("Charts")
        ch.sheet_view.showGridLines = False
        ch.merge_cells("A1:H1")
        ch["A1"] = "BUDGET VISUALS"
        ch["A1"].fill = dark_fill
        ch["A1"].font = title_font
        ch["A1"].alignment = center
        ch["A2"] = "Category"
        ch["B2"] = "Amount"
        ch["C2"] = "Allocation %"

        for cell in ch[2]:
            cell.fill = mid_fill
            cell.font = header_font
            cell.border = border

        for idx, (_, row) in enumerate(df.iterrows(), start=3):
            ch[f"A{idx}"] = row["Category"]
            ch[f"B{idx}"] = row["Amount"]
            ch[f"C{idx}"] = row["Allocation %"] / 100
            ch[f"B{idx}"].number_format = f'"{currency_symbol}" #,##0.00'
            ch[f"C{idx}"].number_format = "0.00%"

        pie = PieChart()
        pie_labels = Reference(ch, min_col=1, min_row=3, max_row=2 + len(df))
        pie_data = Reference(ch, min_col=2, min_row=2, max_row=2 + len(df))
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(pie_labels)
        pie.title = "Budget Distribution"
        pie.height = 9
        pie.width = 12
        ch.add_chart(pie, "E3")

        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Allocation by Category"
        bar.y_axis.title = f"Amount ({currency_code})"
        bar.x_axis.title = "Category"
        bar_data = Reference(ch, min_col=2, min_row=2, max_row=2 + len(df))
        bar_labels = Reference(ch, min_col=1, min_row=3, max_row=2 + len(df))
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(bar_labels)
        bar.height = 9
        bar.width = 14
        ch.add_chart(bar, "E20")

        if not boq_df.empty:
            boq_ws = wb["BOQ Details"]
            boq_ws.sheet_view.showGridLines = False
            boq_has_source = "BOQ Source" in list(boq_ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
            if boq_has_source:
                boq_widths = {"A": 28, "B": 36, "C": 14, "D": 16, "E": 22, "F": 24, "G": 18}
                merge_range = "A1:G1"
            else:
                boq_widths = {"A": 36, "B": 14, "C": 16, "D": 22, "E": 24, "F": 18}
                merge_range = "A1:F1"
            for col, width in boq_widths.items():
                boq_ws.column_dimensions[col].width = width

            boq_ws.merge_cells(merge_range)
            boq_ws["A1"] = "UPLOADED BOQ DETAILS"
            boq_ws["A1"].fill = dark_fill
            boq_ws["A1"].font = title_font
            boq_ws["A1"].alignment = center

            boq_ws["A3"] = "BOQ Total Cost"
            boq_ws["B3"] = boq_summary["total_cost"]
            boq_ws["C3"] = "BOQ Item Count"
            boq_ws["D3"] = boq_summary["total_items"]
            boq_ws["E3"] = "Highest Item"
            boq_ws["F3"] = boq_summary["most_expensive_item"]

            boq_ws["A4"] = "Highest Item Cost"
            boq_ws["B4"] = boq_summary["most_expensive_cost"]
            for cell in ["A3", "C3", "E3", "A4"]:
                boq_ws[cell].fill = soft_fill
                boq_ws[cell].font = header_font
                boq_ws[cell].border = border
            for cell in ["B3", "D3", "F3", "B4"]:
                boq_ws[cell].border = border

            boq_ws["B3"].number_format = f'"{currency_symbol}" #,##0.00'
            boq_ws["B4"].number_format = f'"{currency_symbol}" #,##0.00'

            for cell in boq_ws[8]:
                cell.fill = mid_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

            for row in range(9, 9 + len(boq_df)):
                if boq_has_source:
                    boq_ws[f"C{row}"].number_format = "0.00"
                    boq_ws[f"D{row}"].number_format = f'"{currency_symbol}" #,##0.00'
                    boq_ws[f"G{row}"].number_format = f'"{currency_symbol}" #,##0.00'
                    border_cols = "ABCDEFG"
                else:
                    boq_ws[f"B{row}"].number_format = "0.00"
                    boq_ws[f"C{row}"].number_format = f'"{currency_symbol}" #,##0.00'
                    boq_ws[f"F{row}"].number_format = f'"{currency_symbol}" #,##0.00'
                    border_cols = "ABCDEF"
                for col in border_cols:
                    boq_ws[f"{col}{row}"].border = border
            boq_ws.freeze_panes = "A8"

        if not boq_compare_df.empty:
            cmp_ws = wb["BOQ Comparison"]
            cmp_ws.sheet_view.showGridLines = False
            for col, width in {"A": 28, "B": 20, "C": 20, "D": 20}.items():
                cmp_ws.column_dimensions[col].width = width

            cmp_ws.merge_cells("A1:D1")
            cmp_ws["A1"] = "PLANNED VS ACTUAL BOQ COMPARISON"
            cmp_ws["A1"].fill = dark_fill
            cmp_ws["A1"].font = title_font
            cmp_ws["A1"].alignment = center

            cmp_ws["A3"] = "Project Budget"
            cmp_ws["B3"] = total_budget
            cmp_ws["C3"] = "Allocation Status"
            cmp_ws["D3"] = allocation_status
            for cell in ["A3", "C3"]:
                cmp_ws[cell].fill = soft_fill
                cmp_ws[cell].font = header_font
                cmp_ws[cell].border = border
            for cell in ["B3", "D3"]:
                cmp_ws[cell].border = border
            cmp_ws["B3"].number_format = f'"{currency_symbol}" #,##0.00'

            for cell in cmp_ws[7]:
                cell.fill = mid_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

            for row in range(8, 8 + len(boq_compare_df)):
                for col in "ABCD":
                    cmp_ws[f"{col}{row}"].border = border
                for col in "BCD":
                    cmp_ws[f"{col}{row}"].number_format = f'"{currency_symbol}" #,##0.00'
            cmp_ws.freeze_panes = "A7"

            ch["E38"] = "BOQ Category"
            ch["F38"] = "Planned Budget"
            ch["G38"] = "Actual BOQ Cost"
            for cell in ch[38]:
                if cell.column <= 7:
                    cell.fill = mid_fill
                    cell.font = header_font
                    cell.border = border

            for idx, (_, row) in enumerate(boq_compare_df.iterrows(), start=39):
                ch[f"E{idx}"] = row["Category"]
                ch[f"F{idx}"] = row["Planned Budget"]
                ch[f"G{idx}"] = row["Actual BOQ Cost"]
                ch[f"F{idx}"].number_format = f'"{currency_symbol}" #,##0.00'
                ch[f"G{idx}"].number_format = f'"{currency_symbol}" #,##0.00'

            boq_bar = BarChart()
            boq_bar.type = "col"
            boq_bar.style = 11
            boq_bar.title = "Planned vs Actual BOQ"
            boq_bar.y_axis.title = f"Amount ({currency_code})"
            boq_bar.x_axis.title = "Category"
            boq_data = Reference(ch, min_col=6, min_row=38, max_col=7, max_row=38 + len(boq_compare_df))
            boq_labels = Reference(ch, min_col=5, min_row=39, max_row=38 + len(boq_compare_df))
            boq_bar.add_data(boq_data, titles_from_data=True)
            boq_bar.set_categories(boq_labels)
            boq_bar.height = 10
            boq_bar.width = 14
            ch.add_chart(boq_bar, "E40")

    output.seek(0)
    return output


def create_comparison_excel_report(
    selected_projects_df: pd.DataFrame,
    allocation_history_df: pd.DataFrame,
    comparison_history_df: pd.DataFrame,
    boq_history_df: pd.DataFrame,
    risk_summary_df: pd.DataFrame,
    allocation_pivot_df: pd.DataFrame,
    boq_pivot_df: pd.DataFrame,
    currency_symbol: str,
) -> BytesIO:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wb = writer.book

        dark_fill = PatternFill("solid", fgColor="1F4E78")
        mid_fill = PatternFill("solid", fgColor="D9EAF7")
        soft_fill = PatternFill("solid", fgColor="EDF4FB")
        green_fill = PatternFill("solid", fgColor="E2F0D9")
        red_fill = PatternFill("solid", fgColor="FDE9E7")

        title_font = Font(color="FFFFFF", bold=True, size=15)
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin = Side(style="thin", color="B7B7B7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        overview_df = build_project_overview_dataframe(selected_projects_df)
        if not overview_df.empty:
            overview_df.to_excel(writer, sheet_name="Project Overview", index=False, startrow=3)
            ws = writer.book["Project Overview"]
            ws.sheet_view.showGridLines = False
            ws.merge_cells("A1:J1")
            ws["A1"] = "SELECTED PROJECT COMPARISON OVERVIEW"
            ws["A1"].fill = dark_fill
            ws["A1"].font = title_font
            ws["A1"].alignment = center
            for cell in ws[4]:
                cell.fill = mid_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            for row in ws.iter_rows(min_row=5, max_row=4 + len(overview_df), min_col=1, max_col=len(overview_df.columns)):
                for cell in row:
                    cell.border = border
                    cell.alignment = left
            money_cols = {"F", "G", "H"}
            for row_idx in range(5, 5 + len(overview_df)):
                for col in money_cols:
                    ws[f"{col}{row_idx}"].number_format = f'"{currency_symbol}" #,##0.00'
            widths = {"A": 10, "B": 28, "C": 22, "D": 16, "E": 12, "F": 18, "G": 18, "H": 18, "I": 16, "J": 14}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            ws.freeze_panes = "A4"

        if not allocation_history_df.empty:
            allocation_export_df = allocation_history_df.copy()
            allocation_export_df.to_excel(writer, sheet_name="Allocation Raw", index=False, startrow=3)
            ws = writer.book["Allocation Raw"]
            ws.sheet_view.showGridLines = False
            ws.merge_cells("A1:E1")
            ws["A1"] = "SELECTED PROJECT ALLOCATION DETAILS"
            ws["A1"].fill = dark_fill
            ws["A1"].font = title_font
            ws["A1"].alignment = center
            for cell in ws[4]:
                cell.fill = mid_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            for row in ws.iter_rows(min_row=5, max_row=4 + len(allocation_export_df), min_col=1, max_col=len(allocation_export_df.columns)):
                for cell in row:
                    cell.border = border
            for row_idx in range(5, 5 + len(allocation_export_df)):
                ws[f"D{row_idx}"].number_format = "0.00"
                ws[f"E{row_idx}"].number_format = f'"{currency_symbol}" #,##0.00'
            for col, width in {"A": 12, "B": 26, "C": 24, "D": 16, "E": 18}.items():
                ws.column_dimensions[col].width = width
            ws.freeze_panes = "A4"

        if not allocation_pivot_df.empty:
            allocation_pivot_export = allocation_pivot_df.reset_index()
            allocation_pivot_export.to_excel(writer, sheet_name="Allocation Comparison", index=False, startrow=3)
            ws = writer.book["Allocation Comparison"]
            ws.sheet_view.showGridLines = False
            end_col = chr(64 + min(len(allocation_pivot_export.columns), 26))
            ws.merge_cells(f"A1:{end_col}1")
            ws["A1"] = "ALLOCATION AMOUNT COMPARISON"
            ws["A1"].fill = dark_fill
            ws["A1"].font = title_font
            ws["A1"].alignment = center
            for cell in ws[4]:
                cell.fill = mid_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            for row in ws.iter_rows(min_row=5, max_row=4 + len(allocation_pivot_export), min_col=1, max_col=len(allocation_pivot_export.columns)):
                for cell in row:
                    cell.border = border
            for row_idx in range(5, 5 + len(allocation_pivot_export)):
                for col_idx in range(2, len(allocation_pivot_export.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = f'"{currency_symbol}" #,##0.00'
            for col_idx in range(1, len(allocation_pivot_export.columns) + 1):
                ws.column_dimensions[chr(64 + col_idx)].width = 18 if col_idx > 1 else 24
            ws.freeze_panes = "A4"

        if not comparison_history_df.empty:
            comparison_export_df = comparison_history_df.copy()
            comparison_export_df.to_excel(writer, sheet_name="Planned vs Actual Raw", index=False, startrow=3)
            ws = writer.book["Planned vs Actual Raw"]
            ws.sheet_view.showGridLines = False
            ws.merge_cells("A1:G1")
            ws["A1"] = "PLANNED VS ACTUAL COMPARISON DETAILS"
            ws["A1"].fill = dark_fill
            ws["A1"].font = title_font
            ws["A1"].alignment = center
            for cell in ws[4]:
                cell.fill = mid_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            for row in ws.iter_rows(min_row=5, max_row=4 + len(comparison_export_df), min_col=1, max_col=len(comparison_export_df.columns)):
                for cell in row:
                    cell.border = border
            for row_idx in range(5, 5 + len(comparison_export_df)):
                for col in ["D", "E", "F"]:
                    ws[f"{col}{row_idx}"].number_format = f'"{currency_symbol}" #,##0.00'
                ws[f"G{row_idx}"].number_format = "0.00"
            for col, width in {"A": 12, "B": 26, "C": 24, "D": 18, "E": 18, "F": 18, "G": 16}.items():
                ws.column_dimensions[col].width = width
            ws.freeze_panes = "A4"

        if not boq_pivot_df.empty:
            boq_pivot_export = boq_pivot_df.reset_index()
            boq_pivot_export.to_excel(writer, sheet_name="BOQ Actual Comparison", index=False, startrow=3)
            ws = writer.book["BOQ Actual Comparison"]
            ws.sheet_view.showGridLines = False
            end_col = chr(64 + min(len(boq_pivot_export.columns), 26))
            ws.merge_cells(f"A1:{end_col}1")
            ws["A1"] = "ACTUAL BOQ CATEGORY COMPARISON"
            ws["A1"].fill = dark_fill
            ws["A1"].font = title_font
            ws["A1"].alignment = center
            for cell in ws[4]:
                cell.fill = mid_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            for row in ws.iter_rows(min_row=5, max_row=4 + len(boq_pivot_export), min_col=1, max_col=len(boq_pivot_export.columns)):
                for cell in row:
                    cell.border = border
            for row_idx in range(5, 5 + len(boq_pivot_export)):
                for col_idx in range(2, len(boq_pivot_export.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = f'"{currency_symbol}" #,##0.00'
            for col_idx in range(1, len(boq_pivot_export.columns) + 1):
                ws.column_dimensions[chr(64 + col_idx)].width = 18 if col_idx > 1 else 24
            ws.freeze_panes = "A4"

        if not risk_summary_df.empty:
            risk_summary_df.to_excel(writer, sheet_name="Risk Summary", index=False, startrow=3)
            ws = writer.book["Risk Summary"]
            ws.sheet_view.showGridLines = False
            ws.merge_cells("A1:E1")
            ws["A1"] = "ENGINEERING RISK SUMMARY"
            ws["A1"].fill = dark_fill
            ws["A1"].font = title_font
            ws["A1"].alignment = center
            for cell in ws[4]:
                cell.fill = mid_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            for row in ws.iter_rows(min_row=5, max_row=4 + len(risk_summary_df), min_col=1, max_col=len(risk_summary_df.columns)):
                for cell in row:
                    cell.border = border
                    cell.alignment = wrap_left if cell.column == 5 else center
            for row_idx in range(5, 5 + len(risk_summary_df)):
                ws[f"C{row_idx}"].number_format = "0.00"
                ws[f"D{row_idx}"].number_format = "0.00"
                rating = str(ws[f"B{row_idx}"].value)
                fill = green_fill if "Low" in rating else red_fill if "High" in rating else soft_fill
                for col in "ABCD":
                    ws[f"{col}{row_idx}"].fill = fill
            for col, width in {"A": 12, "B": 18, "C": 14, "D": 18, "E": 60}.items():
                ws.column_dimensions[col].width = width
            ws.freeze_panes = "A4"

        if not selected_projects_df.empty:
            notes_ws = wb.create_sheet("Comparison Notes")
            notes_ws.sheet_view.showGridLines = False
            notes_ws.merge_cells("A1:B1")
            notes_ws["A1"] = "V18 COMPARISON EXPORT NOTES"
            notes_ws["A1"].fill = dark_fill
            notes_ws["A1"].font = title_font
            notes_ws["A1"].alignment = center
            notes = [
                ("Selection Scope", "This workbook only includes projects selected in the comparison dashboard."),
                ("Project Count", str(len(selected_projects_df))),
                ("Generated Sheets", "Overview, allocation comparison, planned vs actual, BOQ actual comparison, and risk summary."),
                ("Use Case", "Use this workbook to compare selected historical jobs without mixing them into the single-project report export."),
            ]
            row = 3
            for label, value in notes:
                notes_ws[f"A{row}"] = label
                notes_ws[f"B{row}"] = value
                notes_ws[f"A{row}"].fill = mid_fill
                notes_ws[f"A{row}"].font = header_font
                notes_ws[f"A{row}"].border = border
                notes_ws[f"B{row}"].border = border
                notes_ws[f"A{row}"].alignment = left
                notes_ws[f"B{row}"].alignment = wrap_left
                row += 1
            notes_ws.column_dimensions["A"].width = 24
            notes_ws.column_dimensions["B"].width = 90

    output.seek(0)
    return output


def format_currency(value: float, currency_symbol: str) -> str:
    return f"{currency_symbol} {value:,.2f}"


def render_alert_box(message: str) -> None:
    if message.startswith("✅"):
        st.success(message)
    elif message.startswith("⚠️"):
        st.warning(message)
    elif message.startswith("❌"):
        st.error(message)
    else:
        st.info(message)


DB_PATH = "project_cost_intelligence.db"


def init_database(db_path: str = DB_PATH) -> None:
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("PRAGMA foreign_keys = ON")

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS projects (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        project_name TEXT NOT NULL,
        project_type TEXT,
        project_reference TEXT,
        client_name TEXT,
        company_name TEXT,
        currency_code TEXT,
        author_name TEXT,
        report_date TEXT,
        total_budget REAL,
        total_allocation_pct REAL,
        allocation_status TEXT,
        boq_total_cost REAL,
        boq_total_items INTEGER,
        most_expensive_item TEXT,
        most_expensive_cost REAL,
        total_difference REAL,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS project_allocations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        project_id INTEGER NOT NULL,
        category TEXT NOT NULL,
        allocation_pct REAL,
        allocation_amount REAL,
        FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS project_boq_categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        project_id INTEGER NOT NULL,
        category TEXT NOT NULL,
        actual_cost REAL,
        actual_pct REAL,
        FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS project_comparison (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        project_id INTEGER NOT NULL,
        category TEXT NOT NULL,
        planned_budget REAL,
        actual_boq_cost REAL,
        difference REAL,
        variance_pct REAL,
        FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
    )
    """)

    conn.commit()
    conn.close()


def save_project_snapshot(
    project_name: str,
    project_type: str,
    project_reference: str,
    client_name: str,
    company_name: str,
    currency_code: str,
    author_name: str,
    report_date: str,
    total_budget: float,
    total_pct: float,
    boq_summary: dict,
    budget_df: pd.DataFrame,
    boq_category_summary: pd.Series,
    boq_compare_df: pd.DataFrame,
    db_path: str = DB_PATH,
) -> int:
    allocation_status = get_allocation_status(total_pct)
    total_difference = float(boq_compare_df["Difference"].sum()) if not boq_compare_df.empty else 0.0

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("PRAGMA foreign_keys = ON")

    cursor.execute(
        """
        INSERT INTO projects (
            project_name,
            project_type,
            project_reference,
            client_name,
            company_name,
            currency_code,
            author_name,
            report_date,
            total_budget,
            total_allocation_pct,
            allocation_status,
            boq_total_cost,
            boq_total_items,
            most_expensive_item,
            most_expensive_cost,
            total_difference
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            project_name,
            project_type,
            project_reference,
            client_name,
            company_name,
            currency_code,
            author_name,
            report_date,
            float(total_budget),
            float(total_pct),
            allocation_status,
            float(boq_summary.get("total_cost", 0.0)),
            int(boq_summary.get("total_items", 0)),
            str(boq_summary.get("most_expensive_item", "N/A")),
            float(boq_summary.get("most_expensive_cost", 0.0)),
            total_difference,
        ),
    )
    project_id = cursor.lastrowid

    allocation_rows = [
        (
            project_id,
            str(row["Category"]),
            float(row["Allocation %"]),
            float(row["Amount"]),
        )
        for _, row in budget_df.iterrows()
    ]
    cursor.executemany(
        """
        INSERT INTO project_allocations (
            project_id, category, allocation_pct, allocation_amount
        )
        VALUES (?, ?, ?, ?)
        """,
        allocation_rows,
    )

    boq_rows = []
    total_boq_cost = float(boq_summary.get("total_cost", 0.0))
    for category, actual_cost in boq_category_summary.items():
        actual_pct = (float(actual_cost) / total_boq_cost * 100) if total_boq_cost else 0.0
        boq_rows.append((project_id, str(category), float(actual_cost), float(actual_pct)))

    if boq_rows:
        cursor.executemany(
            """
            INSERT INTO project_boq_categories (
                project_id, category, actual_cost, actual_pct
            )
            VALUES (?, ?, ?, ?)
            """,
            boq_rows,
        )

    comparison_rows = []
    if not boq_compare_df.empty:
        for _, row in boq_compare_df.iterrows():
            planned = float(row["Planned Budget"])
            actual = float(row["Actual BOQ Cost"])
            difference = float(row["Difference"])
            variance_pct = (difference / planned * 100) if planned else 0.0
            comparison_rows.append(
                (
                    project_id,
                    str(row["Category"]),
                    planned,
                    actual,
                    difference,
                    variance_pct,
                )
            )

    if comparison_rows:
        cursor.executemany(
            """
            INSERT INTO project_comparison (
                project_id, category, planned_budget, actual_boq_cost, difference, variance_pct
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            comparison_rows,
        )

    conn.commit()
    conn.close()
    return project_id


def load_saved_projects(db_path: str = DB_PATH) -> pd.DataFrame:
    conn = sqlite3.connect(db_path)
    query = """
    SELECT
        id,
        project_name,
        project_type,
        project_reference,
        client_name,
        company_name,
        currency_code,
        report_date,
        total_budget,
        boq_total_cost,
        total_allocation_pct,
        allocation_status,
        boq_total_items,
        most_expensive_item,
        total_difference,
        created_at
    FROM projects
    ORDER BY id DESC
    """
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df


def delete_project_snapshot(project_id: int, db_path: str = DB_PATH) -> None:
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("PRAGMA foreign_keys = ON")
    cursor.execute("DELETE FROM project_allocations WHERE project_id = ?", (project_id,))
    cursor.execute("DELETE FROM project_boq_categories WHERE project_id = ?", (project_id,))
    cursor.execute("DELETE FROM project_comparison WHERE project_id = ?", (project_id,))
    cursor.execute("DELETE FROM projects WHERE id = ?", (project_id,))
    conn.commit()
    conn.close()


def load_project_allocations(project_ids: list[int], db_path: str = DB_PATH) -> pd.DataFrame:
    if not project_ids:
        return pd.DataFrame()
    placeholders = ",".join(["?"] * len(project_ids))
    conn = sqlite3.connect(db_path)
    query = f"""
    SELECT
        p.id AS project_id,
        p.project_name,
        a.category,
        a.allocation_pct,
        a.allocation_amount
    FROM project_allocations a
    JOIN projects p ON p.id = a.project_id
    WHERE a.project_id IN ({placeholders})
    ORDER BY p.id DESC, a.id ASC
    """
    df = pd.read_sql_query(query, conn, params=project_ids)
    conn.close()
    return df


def load_project_comparisons(project_ids: list[int], db_path: str = DB_PATH) -> pd.DataFrame:
    if not project_ids:
        return pd.DataFrame()
    placeholders = ",".join(["?"] * len(project_ids))
    conn = sqlite3.connect(db_path)
    query = f"""
    SELECT
        p.id AS project_id,
        p.project_name,
        c.category,
        c.planned_budget,
        c.actual_boq_cost,
        c.difference,
        c.variance_pct
    FROM project_comparison c
    JOIN projects p ON p.id = c.project_id
    WHERE c.project_id IN ({placeholders})
    ORDER BY p.id DESC, c.id ASC
    """
    df = pd.read_sql_query(query, conn, params=project_ids)
    conn.close()
    return df


def load_project_boq_categories(project_ids: list[int], db_path: str = DB_PATH) -> pd.DataFrame:
    if not project_ids:
        return pd.DataFrame()
    placeholders = ",".join(["?"] * len(project_ids))
    conn = sqlite3.connect(db_path)
    query = f"""
    SELECT
        p.id AS project_id,
        p.project_name,
        b.category,
        b.actual_cost,
        b.actual_pct
    FROM project_boq_categories b
    JOIN projects p ON p.id = b.project_id
    WHERE b.project_id IN ({placeholders})
    ORDER BY p.id DESC, b.id ASC
    """
    df = pd.read_sql_query(query, conn, params=project_ids)
    conn.close()
    return df


def build_historical_overview(saved_projects_df: pd.DataFrame) -> pd.DataFrame:
    if saved_projects_df.empty:
        return pd.DataFrame()

    overview_rows = [
        {"Metric": "Average Total Budget", "Value": float(saved_projects_df["total_budget"].mean())},
        {"Metric": "Average BOQ Total Cost", "Value": float(saved_projects_df["boq_total_cost"].mean())},
        {"Metric": "Average BOQ Item Count", "Value": float(saved_projects_df["boq_total_items"].mean())},
        {"Metric": "Average Budget Variance", "Value": float((saved_projects_df["total_budget"] - saved_projects_df["boq_total_cost"]).mean())},
        {"Metric": "Average Allocation %", "Value": float(saved_projects_df["total_allocation_pct"].mean())},
    ]
    return pd.DataFrame(overview_rows)


def build_historical_allocation_benchmark(current_budget_df: pd.DataFrame, allocation_history_df: pd.DataFrame) -> pd.DataFrame:
    if current_budget_df.empty or allocation_history_df.empty:
        return pd.DataFrame()

    historical = (
        allocation_history_df.groupby("category", as_index=False)
        .agg(
            historical_avg_allocation_pct=("allocation_pct", "mean"),
            historical_avg_allocation_amount=("allocation_amount", "mean"),
        )
    )

    current = current_budget_df.rename(
        columns={
            "Category": "category",
            "Allocation %": "current_allocation_pct",
            "Amount": "current_allocation_amount",
        }
    )

    benchmark_df = current.merge(historical, on="category", how="left")
    benchmark_df["allocation_pct_variance"] = (
        benchmark_df["current_allocation_pct"] - benchmark_df["historical_avg_allocation_pct"]
    )
    benchmark_df["allocation_amount_variance"] = (
        benchmark_df["current_allocation_amount"] - benchmark_df["historical_avg_allocation_amount"]
    )
    return benchmark_df.rename(columns={"category": "Category"})


def build_historical_boq_benchmark(boq_summary: dict, boq_category_summary: pd.Series, historical_boq_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    overview_rows = []
    if not historical_boq_df.empty:
        total_cost_history = historical_boq_df.groupby("project_id", as_index=False)["actual_cost"].sum()
        historical_avg_boq_total = float(total_cost_history["actual_cost"].mean()) if not total_cost_history.empty else 0.0
        current_boq_total = float(boq_summary.get("total_cost", 0.0))
        overview_rows.append(
            {
                "Metric": "BOQ Total Cost",
                "Current": current_boq_total,
                "Historical Average": historical_avg_boq_total,
                "Variance": current_boq_total - historical_avg_boq_total,
            }
        )

    overview_df = pd.DataFrame(overview_rows)

    if boq_category_summary.empty or historical_boq_df.empty:
        return overview_df, pd.DataFrame()

    current_total = float(boq_summary.get("total_cost", 0.0))
    current_df = boq_category_summary.reset_index()
    current_df.columns = ["Category", "current_actual_cost"]
    current_df["current_actual_pct"] = current_df["current_actual_cost"].apply(
        lambda value: (float(value) / current_total * 100) if current_total else 0.0
    )

    historical_df = (
        historical_boq_df.groupby("category", as_index=False)
        .agg(
            historical_avg_actual_cost=("actual_cost", "mean"),
            historical_avg_actual_pct=("actual_pct", "mean"),
        )
        .rename(columns={"category": "Category"})
    )

    category_df = current_df.merge(historical_df, on="Category", how="left")
    category_df["actual_cost_variance"] = (
        category_df["current_actual_cost"] - category_df["historical_avg_actual_cost"]
    )
    category_df["actual_pct_variance"] = (
        category_df["current_actual_pct"] - category_df["historical_avg_actual_pct"]
    )
    return overview_df, category_df


def generate_historical_cost_insights(
    total_budget: float,
    current_budget_df: pd.DataFrame,
    boq_df: pd.DataFrame,
    boq_summary: dict,
    saved_projects_df: pd.DataFrame,
    allocation_benchmark_df: pd.DataFrame,
    boq_benchmark_df: pd.DataFrame,
) -> list[str]:
    insights = []

    if saved_projects_df.empty:
        return ["ℹ️ Historical cost intelligence will activate after you save project snapshots."]

    historical_avg_budget = float(saved_projects_df["total_budget"].mean()) if not saved_projects_df.empty else 0.0
    current_boq_total = float(boq_summary.get("total_cost", 0.0))
    current_variance = float(total_budget - current_boq_total) if current_boq_total else 0.0
    historical_avg_variance = float((saved_projects_df["total_budget"] - saved_projects_df["boq_total_cost"]).mean())

    if historical_avg_budget:
        budget_delta_pct = ((total_budget - historical_avg_budget) / historical_avg_budget) * 100
        if budget_delta_pct > 15:
            insights.append(f"⚠️ Current budget is {budget_delta_pct:.1f}% above the historical average budget.")
        elif budget_delta_pct < -15:
            insights.append(f"⚠️ Current budget is {abs(budget_delta_pct):.1f}% below the historical average budget.")
        else:
            insights.append("✅ Current budget is broadly aligned with saved project history.")

    if current_boq_total:
        if historical_avg_variance:
            variance_gap = current_variance - historical_avg_variance
            if variance_gap < -0.01:
                insights.append("⚠️ Current budget variance is weaker than the historical average, so BOQ pressure is higher than usual.")
            else:
                insights.append("✅ Current budget variance is holding at or above the historical average range.")

    if not allocation_benchmark_df.empty:
        for category in ["Materials", "Labor", "Transportation", "Contingency", "Company Profit"]:
            row = allocation_benchmark_df[allocation_benchmark_df["Category"] == category]
            if row.empty or pd.isna(row.iloc[0]["historical_avg_allocation_pct"]):
                continue
            diff = float(row.iloc[0]["allocation_pct_variance"])
            if abs(diff) >= 5:
                direction = "above" if diff > 0 else "below"
                insights.append(
                    f"⚠️ {category} allocation is {abs(diff):.1f} percentage points {direction} the historical average."
                )

    if not boq_benchmark_df.empty:
        for category in ["Materials", "Labor", "Transportation"]:
            row = boq_benchmark_df[boq_benchmark_df["Category"] == category]
            if row.empty or pd.isna(row.iloc[0]["historical_avg_actual_pct"]):
                continue
            diff = float(row.iloc[0]["actual_pct_variance"])
            if abs(diff) >= 7:
                direction = "higher" if diff > 0 else "lower"
                insights.append(
                    f"⚠️ Actual {category.lower()} share is running {abs(diff):.1f} percentage points {direction} than historical BOQ averages."
                )

    if not boq_df.empty:
        top_three_share = float(boq_df.nlargest(min(3, len(boq_df)), "Total Cost")["Total Cost"].sum() / current_boq_total * 100) if current_boq_total else 0.0
        if top_three_share >= 60:
            insights.append(f"⚠️ Top cost concentration is high: the top 3 BOQ items consume {top_three_share:.1f}% of total BOQ value.")
        elif top_three_share >= 40:
            insights.append(f"ℹ️ Top 3 BOQ items account for {top_three_share:.1f}% of total BOQ value.")
        else:
            insights.append(f"✅ BOQ cost concentration looks healthy, with top 3 items at {top_three_share:.1f}% of total BOQ value.")

    if not current_budget_df.empty and current_boq_total:
        materials_row = current_budget_df[current_budget_df["Category"] == "Materials"]
        labor_row = current_budget_df[current_budget_df["Category"] == "Labor"]
        if not materials_row.empty and materials_row.iloc[0]["Amount"] < current_boq_total * 0.35:
            insights.append("⚠️ Material budget looks light versus the uploaded BOQ value. Recheck procurement assumptions.")
        if not labor_row.empty and labor_row.iloc[0]["Amount"] < current_boq_total * 0.15:
            insights.append("⚠️ Labor budget looks compressed against BOQ demand. Verify manpower realism.")

    return insights


def analyze_cost_concentration(boq_df: pd.DataFrame, boq_summary: dict) -> dict:
    current_boq_total = float(boq_summary.get("total_cost", 0.0))
    if boq_df.empty or current_boq_total <= 0:
        return {
            "top_item_share": 0.0,
            "top_3_share": 0.0,
            "top_5_share": 0.0,
            "top_items_label": "No BOQ uploaded",
            "risk_level": "Not Available",
        }

    sorted_df = boq_df.sort_values("Total Cost", ascending=False).copy()
    top_item_share = float(sorted_df.head(1)["Total Cost"].sum() / current_boq_total * 100)
    top_3_share = float(sorted_df.head(min(3, len(sorted_df)))["Total Cost"].sum() / current_boq_total * 100)
    top_5_share = float(sorted_df.head(min(5, len(sorted_df)))["Total Cost"].sum() / current_boq_total * 100)
    top_items_label = ", ".join(sorted_df.head(min(3, len(sorted_df)))["Item"].astype(str).tolist())

    if top_3_share >= 60:
        risk_level = "High"
    elif top_3_share >= 40:
        risk_level = "Moderate"
    else:
        risk_level = "Low"

    return {
        "top_item_share": round(top_item_share, 2),
        "top_3_share": round(top_3_share, 2),
        "top_5_share": round(top_5_share, 2),
        "top_items_label": top_items_label,
        "risk_level": risk_level,
    }


def count_historical_deviations(
    allocation_benchmark_df: pd.DataFrame,
    boq_benchmark_df: pd.DataFrame,
    allocation_threshold: float = 5.0,
    boq_threshold: float = 7.0,
) -> tuple[int, pd.DataFrame]:
    deviation_rows = []

    if not allocation_benchmark_df.empty:
        for _, row in allocation_benchmark_df.iterrows():
            historical_value = row.get("historical_avg_allocation_pct")
            diff = row.get("allocation_pct_variance")
            if pd.notna(historical_value) and pd.notna(diff) and abs(float(diff)) >= allocation_threshold:
                deviation_rows.append(
                    {
                        "Area": "Budget Allocation",
                        "Category": row["Category"],
                        "Current %": float(row["current_allocation_pct"]),
                        "Historical Avg %": float(historical_value),
                        "Variance % Points": float(diff),
                    }
                )

    if not boq_benchmark_df.empty:
        for _, row in boq_benchmark_df.iterrows():
            historical_value = row.get("historical_avg_actual_pct")
            diff = row.get("actual_pct_variance")
            if pd.notna(historical_value) and pd.notna(diff) and abs(float(diff)) >= boq_threshold:
                deviation_rows.append(
                    {
                        "Area": "Actual BOQ",
                        "Category": row["Category"],
                        "Current %": float(row["current_actual_pct"]),
                        "Historical Avg %": float(historical_value),
                        "Variance % Points": float(diff),
                    }
                )

    deviation_df = pd.DataFrame(deviation_rows)
    return len(deviation_rows), deviation_df


def calculate_project_risk_score(
    total_pct: float,
    allocations: dict,
    total_budget: float,
    boq_summary: dict,
    boq_df: pd.DataFrame,
    allocation_benchmark_df: pd.DataFrame,
    boq_benchmark_df: pd.DataFrame,
) -> dict:
    score = 0
    reasons = []
    current_boq_total = float(boq_summary.get("total_cost", 0.0))
    concentration = analyze_cost_concentration(boq_df, boq_summary)
    deviation_count, deviation_df = count_historical_deviations(allocation_benchmark_df, boq_benchmark_df)

    if abs(total_pct - 100) > 1e-9:
        score += 2
        reasons.append("Allocation total is not balanced at 100%.")

    profit = float(allocations.get("Company Profit", 0.0))
    contingency = float(allocations.get("Contingency", 0.0))
    materials = float(allocations.get("Materials", 0.0))
    labor = float(allocations.get("Labor", 0.0))
    transport = float(allocations.get("Transportation", 0.0))

    if profit < 8:
        score += 2
        reasons.append("Profit margin is below the practical 8% comfort level.")

    if contingency < 5:
        score += 2
        reasons.append("Contingency reserve is below the recommended 5% buffer.")

    if materials > 60:
        score += 1
        reasons.append("Materials dependency is unusually high.")

    if labor < 20 or labor > 35:
        score += 1
        reasons.append("Labor allocation sits outside the usual working band.")

    if transport > 10:
        score += 1
        reasons.append("Transportation allocation is elevated and may indicate logistics stress.")

    if current_boq_total and current_boq_total > total_budget:
        score += 2
        reasons.append("Uploaded BOQ total exceeds the current project budget.")

    if concentration["top_3_share"] >= 50:
        score += 2
        reasons.append("Top 3 BOQ items dominate total cost concentration.")

    if deviation_count >= 3:
        score += 2
        reasons.append("Several cost categories are materially deviating from historical norms.")
    elif deviation_count >= 1:
        score += 1
        reasons.append("There are notable deviations against historical category averages.")

    if score <= 2:
        rating = "Low Risk"
    elif score <= 5:
        rating = "Moderate Risk"
    else:
        rating = "High Risk"

    return {
        "score": score,
        "rating": rating,
        "reasons": reasons,
        "deviation_count": deviation_count,
        "deviation_df": deviation_df,
        "concentration": concentration,
    }


def generate_engineering_risk_insights(
    risk_result: dict,
    total_budget: float,
    boq_summary: dict,
    saved_projects_df: pd.DataFrame,
    allocation_benchmark_df: pd.DataFrame,
    boq_benchmark_df: pd.DataFrame,
) -> list[str]:
    insights = []
    current_boq_total = float(boq_summary.get("total_cost", 0.0))
    current_variance = float(total_budget - current_boq_total) if current_boq_total else 0.0

    insights.append(f"ℹ️ Risk score is {risk_result['score']}/10, which places this estimate in the {risk_result['rating'].lower()} band.")

    concentration = risk_result["concentration"]
    if concentration["risk_level"] == "High":
        insights.append(
            f"⚠️ Cost concentration is high: top 3 BOQ items represent {concentration['top_3_share']:.1f}% of total BOQ value."
        )
    elif concentration["risk_level"] == "Moderate":
        insights.append(
            f"⚠️ Cost concentration is moderate: top 3 BOQ items represent {concentration['top_3_share']:.1f}% of total BOQ value."
        )
    elif concentration["risk_level"] == "Low":
        insights.append(
            f"✅ Cost concentration is under control, with top 3 BOQ items at {concentration['top_3_share']:.1f}% of BOQ value."
        )

    if current_boq_total:
        if current_variance < 0:
            insights.append("❌ Current BOQ total is above budget, indicating immediate budget stress.")
        elif current_variance / total_budget < 0.05:
            insights.append("⚠️ Current budget headroom is thin compared with the uploaded BOQ total.")
        else:
            insights.append("✅ Budget headroom against the uploaded BOQ remains workable.")

    if not saved_projects_df.empty:
        historical_avg_variance = float((saved_projects_df["total_budget"] - saved_projects_df["boq_total_cost"]).mean())
        if current_boq_total and current_variance < historical_avg_variance:
            insights.append("⚠️ Current budget variance is below the historical average, so execution flexibility is weaker than normal.")
        elif current_boq_total:
            insights.append("✅ Budget variance is tracking at or above the historical average range.")

    if not allocation_benchmark_df.empty:
        tight_profit = allocation_benchmark_df[allocation_benchmark_df["Category"] == "Company Profit"]
        if not tight_profit.empty and pd.notna(tight_profit.iloc[0]["historical_avg_allocation_pct"]):
            diff = float(tight_profit.iloc[0]["allocation_pct_variance"])
            if diff <= -3:
                insights.append("⚠️ Company profit is materially below historical average. Commercial cushion is tightening.")
        contingency_row = allocation_benchmark_df[allocation_benchmark_df["Category"] == "Contingency"]
        if not contingency_row.empty and pd.notna(contingency_row.iloc[0]["historical_avg_allocation_pct"]):
            diff = float(contingency_row.iloc[0]["allocation_pct_variance"])
            if diff <= -2:
                insights.append("⚠️ Contingency is below historical average. Site surprises may hit harder than usual.")

    if not boq_benchmark_df.empty:
        materials_row = boq_benchmark_df[boq_benchmark_df["Category"] == "Materials"]
        labor_row = boq_benchmark_df[boq_benchmark_df["Category"] == "Labor"]
        if not materials_row.empty and pd.notna(materials_row.iloc[0]["historical_avg_actual_pct"]):
            diff = float(materials_row.iloc[0]["actual_pct_variance"])
            if diff >= 7:
                insights.append("⚠️ Material share is running above historical BOQ norms. Procurement pressure may increase.")
        if not labor_row.empty and pd.notna(labor_row.iloc[0]["historical_avg_actual_pct"]):
            diff = float(labor_row.iloc[0]["actual_pct_variance"])
            if diff <= -7:
                insights.append("⚠️ Labor share is below historical BOQ norms. Execution manpower may be underweighted.")

    return insights


def generate_execution_risk_note(
    project_name: str,
    risk_result: dict,
    total_budget: float,
    boq_summary: dict,
    saved_projects_df: pd.DataFrame,
) -> str:
    current_boq_total = float(boq_summary.get("total_cost", 0.0))
    current_variance = float(total_budget - current_boq_total) if current_boq_total else 0.0
    concentration = risk_result["concentration"]

    drivers = []
    if current_boq_total and current_variance < 0:
        drivers.append("BOQ value exceeding the planned budget")
    elif current_boq_total and current_variance / total_budget < 0.05:
        drivers.append("thin budget headroom")

    if concentration["top_3_share"] >= 50:
        drivers.append(f"high cost concentration in the top BOQ items ({concentration['top_3_share']:.1f}% across the top 3 items)")

    if risk_result["deviation_count"] >= 3:
        drivers.append("multiple deviations from historical category averages")

    if not drivers:
        drivers.append("a generally stable balance between allocation, variance, and concentration")

    note = (
        f"{project_name} currently falls in the {risk_result['rating'].lower()} band with a risk score of "
        f"{risk_result['score']}/10. The main drivers are {', '.join(drivers)}. "
    )

    if not saved_projects_df.empty and current_boq_total:
        historical_avg_variance = float((saved_projects_df["total_budget"] - saved_projects_df["boq_total_cost"]).mean())
        if current_variance < historical_avg_variance:
            note += "Compared with saved project history, execution flexibility is softer than usual and deserves closer review. "
        else:
            note += "Compared with saved project history, the project remains within a workable execution range. "

    note += "Use this note as an engineering review prompt before final commercial submission."
    return note


def build_project_overview_dataframe(projects_df: pd.DataFrame) -> pd.DataFrame:
    if projects_df.empty:
        return pd.DataFrame()
    overview = projects_df.copy()
    overview["Budget Variance"] = overview["total_budget"] - overview["boq_total_cost"]
    return overview[[
        "id",
        "project_name",
        "project_type",
        "report_date",
        "currency_code",
        "total_budget",
        "boq_total_cost",
        "Budget Variance",
        "allocation_status",
        "boq_total_items",
    ]]


def render_history_insight(projects_df: pd.DataFrame) -> None:
    if projects_df.empty:
        st.info("Save your first project snapshot to start the comparison history.")
        return

    highest_budget = projects_df.loc[projects_df["total_budget"].idxmax()]
    highest_boq = projects_df.loc[projects_df["boq_total_cost"].idxmax()]
    st.caption(
        f"Highest budget project: **{highest_budget['project_name']}** "
        f"({highest_budget['currency_code']} {highest_budget['total_budget']:,.2f}) · "
        f"Highest BOQ total: **{highest_boq['project_name']}** "
        f"({highest_boq['currency_code']} {highest_boq['boq_total_cost']:,.2f})"
    )


init_database()

st.markdown(
    """
    <style>
    .block-container {
        padding-top: 1.5rem;
        padding-bottom: 2rem;
    }
    .stMetric {
        background: linear-gradient(180deg, rgba(31,78,120,0.07), rgba(217,234,247,0.28));
        border: 1px solid rgba(31,78,120,0.12);
        border-radius: 16px;
        padding: 10px;
    }
    div[data-testid="stDataFrame"] {
        border-radius: 14px;
        overflow: hidden;
        border: 1px solid rgba(0,0,0,0.08);
    }
    .v14-hero {
        padding: 1.15rem 1.25rem;
        border: 1px solid rgba(31,78,120,0.12);
        border-radius: 18px;
        background: linear-gradient(135deg, rgba(217,234,247,0.65), rgba(255,255,255,0.92));
        margin-bottom: 1rem;
    }
    .v14-label {
        font-size: 0.85rem;
        opacity: 0.8;
        margin-bottom: 0.15rem;
    }
    .v14-value {
        font-size: 1.05rem;
        font-weight: 600;
        margin-bottom: 0.8rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

boq_template_df = pd.DataFrame(
    {
        "Item": ["Steel Structure", "Installation Labor", "Transport Charges"],
        "Quantity": [12, 30, 1],
        "Unit Cost": [2500, 180, 3500],
        "Category": ["Materials", "Labor", "Transportation"],
    }
)

with st.sidebar:
    st.header("⚙️ Project Controls")
    st.caption("V19.1 adds BOQ mode selection so uploaded BOQs can be merged into one project or analyzed as separate cases.")

    st.markdown("### Project Details")
    project_name = st.text_input("Project Name", value="New Engineering Project")
    client_name = st.text_input("Client Name", value="Client Name")
    company_name = st.text_input("Company Name", value="TKM")
    project_reference = st.text_input("Project Reference", value="TKM-EST-2026-001")
    logo_path_input = st.text_input("Company Logo Path (optional)", value="")

    author_name = "Abdulla Zahin"
    today = date.today().strftime("%d %B %Y")
    currency_code = st.selectbox("Currency", options=["AED", "USD", "INR"], index=0)
    st.text_input("Author", value=author_name, disabled=True)
    st.text_input("Date", value=today, disabled=True)

    st.markdown("### BOQ Upload")
    st.download_button(
        label="⬇️ Download BOQ Template",
        data=boq_template_df.to_csv(index=False).encode("utf-8"),
        file_name="boq_template.csv",
        mime="text/csv",
        use_container_width=True,
    )
    uploaded_boq_files = st.file_uploader("Upload BOQ File(s)", type=["csv", "xlsx"], accept_multiple_files=True)

    st.markdown("### Budget Setup")
    project_type = st.selectbox("Project Type", options=list(PROJECT_TYPE_SPLITS.keys()), index=0)

    if "last_project_type" not in st.session_state:
        st.session_state.last_project_type = project_type
    if "allocations_state" not in st.session_state:
        reset_allocation_widget_state(project_type)
    elif project_type != st.session_state.last_project_type:
        reset_allocation_widget_state(project_type)

    selected_split = st.session_state.allocations_state

    total_budget = st.number_input(
        "Total Project Budget",
        min_value=0.0,
        value=500000.0,
        step=10000.0,
        format="%.2f",
    )

    st.caption(f"Smart recommendation is loaded for **{project_type}**. You can still tune the percentages below.")
    if st.button("Reset to Project Defaults", use_container_width=True):
        reset_allocation_widget_state(project_type)
        st.rerun()

    allocations = {}
    for category, _ in selected_split.items():
        widget_key = f"alloc_{category}"
        allocations[category] = st.number_input(
            f"{category} (%)",
            min_value=0.0,
            max_value=100.0,
            value=float(st.session_state.allocations_state[category]),
            step=1.0,
            key=widget_key,
        )
        st.session_state.allocations_state[category] = allocations[category]

currency_symbol = CURRENCY_SYMBOLS[currency_code]

boq_df = pd.DataFrame()
boq_category_summary = pd.Series(dtype=float)
boq_summary = {
    "total_cost": 0.0,
    "total_items": 0,
    "most_expensive_item": "N/A",
    "most_expensive_cost": 0.0,
}
boq_compare_df = pd.DataFrame()
selected_boq_names = []
selected_boq_uploads = []
separate_case_results = []
separate_case_summary_df = pd.DataFrame()
uploaded_boq_count = len(uploaded_boq_files) if uploaded_boq_files else 0
boq_processing_mode = "Separate Cases"

if uploaded_boq_files:
    uploaded_boq_names = [file.name for file in uploaded_boq_files]
    default_selected_boq_names = uploaded_boq_names.copy()

    with st.sidebar:
        st.markdown("#### Select BOQs to Include")
        selected_boq_names = st.multiselect(
            "Included BOQ Files",
            options=uploaded_boq_names,
            default=default_selected_boq_names,
            help="All uploaded BOQs are selected by default. Remove any file you do not want to include in the current BOQ analysis.",
        )

        if len(uploaded_boq_files) > 1:
            merge_selected_boqs = st.checkbox(
                "Merge selected BOQs into one project",
                value=False,
                help="Turn this on only when the uploaded BOQs are parts of one project. Leave it off to treat each selected BOQ as a separate case.",
            )
            boq_processing_mode = "Merge as Single Project" if merge_selected_boqs else "Separate Cases"
        else:
            boq_processing_mode = "Merge as Single Project"

    if selected_boq_names:
        selected_boq_uploads = [file for file in uploaded_boq_files if file.name in selected_boq_names]
        merged_boq_dfs = []
        boq_file_errors = []

        for uploaded_file in selected_boq_uploads:
            try:
                raw_boq_df = load_boq_file(uploaded_file)
                cleaned_boq_df = clean_boq_dataframe(raw_boq_df)
                cleaned_boq_df["BOQ Source"] = uploaded_file.name

                case_summary, case_category_summary = analyze_boq_items(cleaned_boq_df)
                separate_case_results.append(
                    {
                        "file_name": uploaded_file.name,
                        "boq_df": cleaned_boq_df,
                        "summary": case_summary,
                        "category_summary": case_category_summary,
                    }
                )

                if boq_processing_mode == "Merge as Single Project":
                    merged_boq_dfs.append(cleaned_boq_df)
            except Exception as exc:
                boq_file_errors.append(f"{uploaded_file.name}: {exc}")

        for error_message in boq_file_errors:
            st.error(f"BOQ upload error: {error_message}")

        separate_case_summary_df = build_separate_case_summary(separate_case_results)

        if boq_processing_mode == "Merge as Single Project" and merged_boq_dfs:
            boq_df = pd.concat(merged_boq_dfs, ignore_index=True)
            boq_summary, boq_category_summary = analyze_boq_items(boq_df)
    else:
        st.warning("Upload detected, but no BOQ files are currently selected for analysis.")


total_pct = sum(allocations.values())
remaining_pct = 100 - total_pct
profit_amount = total_budget * allocations["Company Profit"] / 100
materials_amount = total_budget * allocations["Materials"] / 100
labor_amount = total_budget * allocations["Labor"] / 100
contingency_amount = total_budget * allocations["Contingency"] / 100


df = build_budget_dataframe(total_budget, allocations)
insights = analyze_budget(allocations, total_pct)
report_sections = generate_report_text(
    project_name=project_name,
    project_type=project_type,
    company_name=company_name,
    client_name=client_name,
    project_reference=project_reference,
    currency_symbol=currency_symbol,
    total_budget=total_budget,
    allocations=allocations,
    total_pct=total_pct,
)

if not boq_df.empty:
    boq_compare_df = build_boq_comparison_dataframe(total_budget, allocations, boq_df)

st.markdown("""
<div class='v14-hero'>
    <div style='font-size: 2rem; font-weight: 700;'>🏗️ Project Cost Intelligence System - V19.1</div>
    <div style='font-size: 1rem; opacity: 0.88; margin-top: 0.2rem;'>Professional UI upgrade for smart BOQ analysis, budget planning, export-ready reporting, and cleaner decision support.</div>
</div>
""", unsafe_allow_html=True)

hero_col1, hero_col2, hero_col3, hero_col4 = st.columns([1.4, 1.1, 1.1, 1.1])
with hero_col1:
    st.markdown(f"<div class='v14-label'>Project</div><div class='v14-value'>{project_name}</div>", unsafe_allow_html=True)
    st.markdown(f"<div class='v14-label'>Client</div><div class='v14-value'>{client_name}</div>", unsafe_allow_html=True)
with hero_col2:
    st.markdown(f"<div class='v14-label'>Company</div><div class='v14-value'>{company_name}</div>", unsafe_allow_html=True)
    st.markdown(f"<div class='v14-label'>Project Type</div><div class='v14-value'>{project_type}</div>", unsafe_allow_html=True)
with hero_col3:
    st.markdown(f"<div class='v14-label'>Reference</div><div class='v14-value'>{project_reference}</div>", unsafe_allow_html=True)
    st.markdown(f"<div class='v14-label'>Currency</div><div class='v14-value'>{currency_code}</div>", unsafe_allow_html=True)
with hero_col4:
    st.markdown(f"<div class='v14-label'>Prepared By</div><div class='v14-value'>{author_name}</div>", unsafe_allow_html=True)
    st.markdown(f"<div class='v14-label'>Date</div><div class='v14-value'>{today}</div>", unsafe_allow_html=True)

kpi1, kpi2, kpi3, kpi4, kpi5 = st.columns(5)
kpi1.metric("Total Budget", format_currency(total_budget, currency_symbol))
kpi2.metric("Total Allocation %", f"{total_pct:.2f}%")
kpi3.metric("Unallocated / Excess %", f"{remaining_pct:.2f}%")
kpi4.metric("Estimated Profit", format_currency(profit_amount, currency_symbol), f"{allocations['Company Profit']:.2f}%")
kpi5.metric("BOQ Total Cost", format_currency(boq_summary['total_cost'], currency_symbol) if not boq_df.empty else ("Separate mode" if boq_processing_mode == "Separate Cases" and not separate_case_summary_df.empty else "Not uploaded"))

sum1, sum2, sum3, sum4 = st.columns(4)
sum1.metric("Materials", format_currency(materials_amount, currency_symbol))
sum2.metric("Labor", format_currency(labor_amount, currency_symbol))
sum3.metric("Contingency", format_currency(contingency_amount, currency_symbol))
sum4.metric("BOQ Items / Files", f"{boq_summary['total_items']} / {len(selected_boq_names) if selected_boq_names else 0}" if boq_processing_mode == "Merge as Single Project" else f"{int(separate_case_summary_df['BOQ Items'].sum()) if not separate_case_summary_df.empty else 0} / {len(selected_boq_names) if selected_boq_names else 0}")

st.markdown("## Budget Intelligence")
for item in insights:
    render_alert_box(item)
if uploaded_boq_files and boq_processing_mode == "Merge as Single Project" and boq_df.empty:
    st.warning("BOQ file was uploaded but could not be analyzed. Please check the file structure.")
elif uploaded_boq_files and boq_processing_mode == "Separate Cases" and not separate_case_summary_df.empty:
    st.success("BOQ file(s) uploaded and analyzed in separate case mode.")
    selected_count = len(selected_boq_names) if selected_boq_names else 0
    upload_summary_col1, upload_summary_col2 = st.columns(2)
    upload_summary_col1.caption(f"BOQ files included in current analysis: **{selected_count} / {uploaded_boq_count}**")
    upload_summary_col2.caption(f"Processing mode: **Separate Cases**")
elif not boq_df.empty:
    st.success("BOQ file(s) uploaded and analyzed successfully.")
    selected_count = len(selected_boq_names) if selected_boq_names else 0
    upload_summary_col1, upload_summary_col2 = st.columns(2)
    upload_summary_col1.caption(f"BOQ files included in current analysis: **{selected_count} / {uploaded_boq_count}**")
    upload_summary_col2.caption(f"Merged BOQ items loaded: **{boq_summary['total_items']}**")
else:
    st.info("Upload a BOQ file from the sidebar whenever you want planned vs actual cost intelligence.")

break_col1, break_col2 = st.columns([1.05, 0.95])
with break_col1:
    st.markdown("## Budget Breakdown")
    st.dataframe(df, use_container_width=True, hide_index=True)

with break_col2:
    st.markdown("## Allocation Snapshot")
    fig1, ax1 = plt.subplots(figsize=(6, 5.2))
    ax1.pie(df["Amount"], labels=df["Category"], autopct="%1.1f%%", startangle=90)
    ax1.set_title("Budget Distribution")
    st.pyplot(fig1)
    plt.close(fig1)

st.markdown("## Visual Dashboard")
chart_col1, chart_col2 = st.columns(2)
with chart_col1:
    fig2, ax2 = plt.subplots(figsize=(8, 5))
    ax2.bar(df["Category"], df["Amount"])
    ax2.set_title("Category-wise Budget Amount")
    ax2.set_xlabel("Category")
    ax2.set_ylabel(f"Amount ({currency_code})")
    plt.xticks(rotation=45, ha="right")
    st.pyplot(fig2)
    plt.close(fig2)

with chart_col2:
    top_items_df = boq_df.sort_values("Total Cost", ascending=False).head(10) if not boq_df.empty else pd.DataFrame()
    if not top_items_df.empty:
        fig3, ax3 = plt.subplots(figsize=(8, 5))
        ax3.bar(top_items_df["Item"], top_items_df["Total Cost"])
        ax3.set_title("Top 10 BOQ Cost Drivers")
        ax3.set_xlabel("Item")
        ax3.set_ylabel(f"Amount ({currency_code})")
        plt.xticks(rotation=45, ha="right")
        st.pyplot(fig3)
        plt.close(fig3)
    else:
        st.info("Top BOQ cost driver chart will appear after BOQ upload.")

if uploaded_boq_files and boq_processing_mode == "Separate Cases" and not separate_case_summary_df.empty:
    st.markdown("## BOQ Case Intelligence")
    case_metric1, case_metric2, case_metric3 = st.columns(3)
    case_metric1.metric("Separate BOQ Cases", str(len(separate_case_results)))
    case_metric2.metric("Total Items Across Cases", str(int(separate_case_summary_df["BOQ Items"].sum())))
    case_metric3.metric("Highest Case Total", format_currency(float(separate_case_summary_df["Total Cost"].max()), currency_symbol))

    display_case_summary_df = separate_case_summary_df.copy()
    st.dataframe(display_case_summary_df, use_container_width=True, hide_index=True)

    case_chart_df = separate_case_summary_df.copy()
    if not case_chart_df.empty:
        fig_case, ax_case = plt.subplots(figsize=(8, 5))
        ax_case.bar(case_chart_df["BOQ File"], case_chart_df["Total Cost"])
        ax_case.set_title("Total Cost by Uploaded BOQ Case")
        ax_case.set_xlabel("BOQ File")
        ax_case.set_ylabel(f"Amount ({currency_code})")
        plt.xticks(rotation=45, ha="right")
        st.pyplot(fig_case)
        plt.close(fig_case)

    case_labels = [case["file_name"] for case in separate_case_results]
    selected_case_label = st.selectbox("Select a BOQ case to inspect", options=case_labels, key="separate_case_selector")
    selected_case = next((case for case in separate_case_results if case["file_name"] == selected_case_label), None)

    if selected_case is not None:
        selected_case_df = selected_case["boq_df"]
        selected_case_summary = selected_case["summary"]
        selected_case_category_summary = selected_case["category_summary"]
        case_tab1, case_tab2 = st.tabs(["Case Preview", "Case Category Summary"])

        with case_tab1:
            st.dataframe(
                selected_case_df[["BOQ Source", "Item", "Quantity", "Unit Cost", "Category", "Normalized Category", "Total Cost"]],
                use_container_width=True,
                hide_index=True,
            )

        with case_tab2:
            case_category_df = selected_case_category_summary.reset_index()
            case_category_df.columns = ["Category", "BOQ Amount"]
            st.dataframe(case_category_df, use_container_width=True, hide_index=True)
            st.caption(
                f"Selected case total: {format_currency(float(selected_case_summary['total_cost']), currency_symbol)} · "
                f"Items: {int(selected_case_summary['total_items'])} · "
                f"Highest item: {selected_case_summary['most_expensive_item']}"
            )

    st.info("Separate case mode keeps uploaded BOQs independent. Single-project BOQ comparison, historical BOQ benchmarking, and BOQ-based exports continue to use merge mode only.")

if not boq_df.empty:
    st.markdown("## BOQ Intelligence")
    boq_m1, boq_m2, boq_m3, boq_m4 = st.columns(4)
    boq_m1.metric("BOQ Total Cost", format_currency(boq_summary["total_cost"], currency_symbol))
    boq_m2.metric("BOQ Items", f"{boq_summary['total_items']}")
    boq_m3.metric("Highest BOQ Item", boq_summary["most_expensive_item"])
    boq_m4.metric("Highest BOQ Item Cost", format_currency(boq_summary["most_expensive_cost"], currency_symbol))
    st.caption(f"BOQ files merged into this analysis: **{len(selected_boq_names) if selected_boq_names else 0}**")

    boq_tab1, boq_tab2, boq_tab3 = st.tabs(["BOQ Preview", "Category Summary", "Planned vs Actual"])

    with boq_tab1:
        boq_preview_columns = ["BOQ Source", "Item", "Quantity", "Unit Cost", "Category", "Normalized Category", "Total Cost"] if "BOQ Source" in boq_df.columns else ["Item", "Quantity", "Unit Cost", "Category", "Normalized Category", "Total Cost"]
        st.dataframe(
            boq_df[boq_preview_columns],
            use_container_width=True,
            hide_index=True,
        )

    with boq_tab2:
        boq_category_df = boq_category_summary.reset_index()
        boq_category_df.columns = ["Category", "BOQ Amount"]
        st.dataframe(boq_category_df, use_container_width=True, hide_index=True)

    with boq_tab3:
        st.dataframe(boq_compare_df, use_container_width=True, hide_index=True)
        compare_fig, compare_ax = plt.subplots(figsize=(9, 5))
        x_positions = range(len(boq_compare_df))
        width = 0.38
        compare_ax.bar([x - width / 2 for x in x_positions], boq_compare_df["Planned Budget"], width=width, label="Planned")
        compare_ax.bar([x + width / 2 for x in x_positions], boq_compare_df["Actual BOQ Cost"], width=width, label="Actual BOQ")
        compare_ax.set_title("Planned vs Actual BOQ by Category")
        compare_ax.set_xlabel("Category")
        compare_ax.set_ylabel(f"Amount ({currency_code})")
        compare_ax.set_xticks(list(x_positions))
        compare_ax.set_xticklabels(boq_compare_df["Category"], rotation=45, ha="right")
        compare_ax.legend()
        st.pyplot(compare_fig)
        plt.close(compare_fig)

st.markdown("## Auto Report Writing")
report_help_col1, report_help_col2 = st.columns([2, 3])
with report_help_col1:
    report_option = st.selectbox("Select Report Draft Type", options=list(report_sections.keys()))
with report_help_col2:
    st.caption("These draft sections stay export-ready so they can move straight into proposal notes, internal reviews, or client-facing summaries.")

st.text_area("Generated Draft", value=report_sections[report_option], height=220)
full_report_text = "\n\n".join(f"{title}\n{content}" for title, content in report_sections.items())
st.download_button(
    label="📝 Download Report Draft (.txt)",
    data=full_report_text.encode("utf-8"),
    file_name="Project_Report_Draft.txt",
    mime="text/plain",
)


st.markdown("## 📚 V15/V16 - Multi-Project Comparison")

save_col1, save_col2 = st.columns([1.3, 1.7])
with save_col1:
    if boq_processing_mode == "Separate Cases" and separate_case_results:
        st.caption("Save each selected BOQ as its own project snapshot for future comparisons.")
        if st.button("💾 Save Separate Project Snapshots", use_container_width=True):
            saved_ids = []
            for idx, case in enumerate(separate_case_results, start=1):
                case_file_name = str(case["file_name"])
                case_stem = Path(case_file_name).stem
                case_project_name = f"{project_name} - {case_stem}"
                case_project_reference = f"{project_reference}-{idx:02d}"
                case_boq_df = case["boq_df"]
                case_summary = case["summary"]
                case_category_summary = case["category_summary"]
                case_compare_df = build_boq_comparison_dataframe(total_budget, allocations, case_boq_df)
                saved_project_id = save_project_snapshot(
                    project_name=case_project_name,
                    project_type=project_type,
                    project_reference=case_project_reference,
                    client_name=client_name,
                    company_name=company_name,
                    currency_code=currency_code,
                    author_name=author_name,
                    report_date=today,
                    total_budget=total_budget,
                    total_pct=total_pct,
                    boq_summary=case_summary,
                    budget_df=df,
                    boq_category_summary=case_category_summary,
                    boq_compare_df=case_compare_df,
                )
                saved_ids.append(str(saved_project_id))
            st.success(f"Saved {len(saved_ids)} separate project snapshots successfully! Project IDs: {', '.join(saved_ids)}")
    else:
        st.caption("Save the current analysis as a reusable project snapshot for future comparisons.")
        if st.button("💾 Save Project Snapshot", use_container_width=True):
            saved_project_id = save_project_snapshot(
                project_name=project_name,
                project_type=project_type,
                project_reference=project_reference,
                client_name=client_name,
                company_name=company_name,
                currency_code=currency_code,
                author_name=author_name,
                report_date=today,
                total_budget=total_budget,
                total_pct=total_pct,
                boq_summary=boq_summary,
                budget_df=df,
                boq_category_summary=boq_category_summary,
                boq_compare_df=boq_compare_df,
            )
            st.success(f"Project snapshot saved successfully! Project ID: {saved_project_id}")

with save_col2:
    st.caption("The database file is created automatically by Python and stores all saved project snapshots for V15 comparison.")

saved_projects_df = load_saved_projects()
render_history_insight(saved_projects_df)

history_tab, compare_tab = st.tabs(["Saved Project History", "Project Comparison Dashboard"])

with history_tab:
    if saved_projects_df.empty:
        st.info("No saved projects yet.")
    else:
        history_display_df = build_project_overview_dataframe(saved_projects_df)
        st.dataframe(history_display_df, use_container_width=True, hide_index=True)

        st.markdown("### 🗑 Manage Saved Projects")
        delete_options = {
            f"{row['id']} - {row['project_name']} ({row['report_date']})": int(row["id"])
            for _, row in saved_projects_df.iterrows()
        }
        delete_label = st.selectbox(
            "Select a saved project to delete",
            options=list(delete_options.keys()),
            key="delete_project_select",
        )
        st.warning("Deleting a project will also remove its saved allocations, BOQ categories, and comparison history.")
        confirm_delete = st.checkbox(
            "I understand that this action cannot be undone.",
            key="confirm_project_delete",
        )

        if st.button("Delete Selected Project", type="primary", use_container_width=True):
            if not confirm_delete:
                st.error("Please confirm deletion before removing the project.")
            else:
                project_id_to_delete = delete_options[delete_label]
                deleted_project_name = delete_label.split(" - ", 1)[1]
                delete_project_snapshot(project_id_to_delete)
                st.success(f"Deleted project snapshot: {deleted_project_name}")
                st.rerun()

with compare_tab:
    if saved_projects_df.empty:
        st.info("Save at least 2 projects to unlock the comparison dashboard.")
    else:
        project_labels = [
            f"{row['id']} - {row['project_name']} ({row['report_date']})"
            for _, row in saved_projects_df.iterrows()
        ]
        default_selection = project_labels[: min(2, len(project_labels))]
        selected_labels = st.multiselect(
            "Select projects to compare",
            options=project_labels,
            default=default_selection,
        )
        selected_ids = [int(label.split(" - ")[0]) for label in selected_labels]

        if selected_ids:
            selected_projects_df = saved_projects_df[saved_projects_df["id"].isin(selected_ids)].copy()
            comparison_overview = build_project_overview_dataframe(selected_projects_df)
            st.dataframe(comparison_overview, use_container_width=True, hide_index=True)

            compare_metric1, compare_metric2, compare_metric3 = st.columns(3)
            compare_metric1.metric("Projects Selected", len(selected_projects_df))
            compare_metric2.metric(
                "Highest Budget",
                f"{selected_projects_df.iloc[0]['currency_code']} {selected_projects_df['total_budget'].max():,.2f}",
            )
            compare_metric3.metric(
                "Highest BOQ Cost",
                f"{selected_projects_df.iloc[0]['currency_code']} {selected_projects_df['boq_total_cost'].max():,.2f}",
            )

            comp_chart_col1, comp_chart_col2 = st.columns(2)
            with comp_chart_col1:
                fig_hist_1, ax_hist_1 = plt.subplots(figsize=(8, 5))
                ax_hist_1.bar(selected_projects_df["project_name"], selected_projects_df["total_budget"])
                ax_hist_1.set_title("Total Budget by Project")
                ax_hist_1.set_xlabel("Project")
                ax_hist_1.set_ylabel(f"Amount ({selected_projects_df.iloc[0]['currency_code']})")
                plt.xticks(rotation=45, ha="right")
                st.pyplot(fig_hist_1)
                plt.close(fig_hist_1)

            with comp_chart_col2:
                fig_hist_2, ax_hist_2 = plt.subplots(figsize=(8, 5))
                ax_hist_2.bar(selected_projects_df["project_name"], selected_projects_df["boq_total_cost"])
                ax_hist_2.set_title("BOQ Total Cost by Project")
                ax_hist_2.set_xlabel("Project")
                ax_hist_2.set_ylabel(f"Amount ({selected_projects_df.iloc[0]['currency_code']})")
                plt.xticks(rotation=45, ha="right")
                st.pyplot(fig_hist_2)
                plt.close(fig_hist_2)

            allocation_history_df = load_project_allocations(selected_ids)
            if not allocation_history_df.empty:
                st.markdown("### Allocation Comparison")
                st.dataframe(allocation_history_df, use_container_width=True, hide_index=True)

                allocation_history_df["project_label"] = (
                    allocation_history_df["project_name"].astype(str)
                    + " (#"
                    + allocation_history_df["project_id"].astype(str)
                    + ")"
                )
                pivot_alloc = allocation_history_df.pivot_table(
                    index="category",
                    columns="project_label",
                    values="allocation_amount",
                    aggfunc="sum",
                    fill_value=0.0,
                )
                fig_alloc, ax_alloc = plt.subplots(figsize=(9, 5))
                pivot_alloc.plot(kind="bar", ax=ax_alloc)
                ax_alloc.set_title("Allocation Amount by Category")
                ax_alloc.set_xlabel("Category")
                ax_alloc.set_ylabel("Amount")
                plt.xticks(rotation=45, ha="right")
                st.pyplot(fig_alloc)
                plt.close(fig_alloc)

            comparison_history_df = load_project_comparisons(selected_ids)
            if not comparison_history_df.empty:
                st.markdown("### Planned vs Actual Comparison")
                st.dataframe(comparison_history_df, use_container_width=True, hide_index=True)

            selected_boq_history_df = load_project_boq_categories(selected_ids)
            allocation_pivot_df = pd.DataFrame()
            boq_pivot_df = pd.DataFrame()
            risk_summary_rows = []

            if not allocation_history_df.empty:
                allocation_pivot_df = allocation_history_df.pivot_table(
                    index="category",
                    columns="project_label",
                    values="allocation_amount",
                    aggfunc="sum",
                    fill_value=0.0,
                )

            if not selected_boq_history_df.empty:
                selected_boq_history_df["project_label"] = (
                    selected_boq_history_df["project_name"].astype(str)
                    + " (#"
                    + selected_boq_history_df["project_id"].astype(str)
                    + ")"
                )
                boq_pivot_df = selected_boq_history_df.pivot_table(
                    index="category",
                    columns="project_label",
                    values="actual_cost",
                    aggfunc="sum",
                    fill_value=0.0,
                )
                st.markdown("### Actual BOQ Category Comparison")
                st.dataframe(selected_boq_history_df, use_container_width=True, hide_index=True)

            for _, project_row in selected_projects_df.iterrows():
                project_id = int(project_row["id"])
                project_alloc_df = allocation_history_df[allocation_history_df["project_id"] == project_id].copy() if not allocation_history_df.empty else pd.DataFrame()
                project_cmp_df = comparison_history_df[comparison_history_df["project_id"] == project_id].copy() if not comparison_history_df.empty else pd.DataFrame()
                project_boq_hist_df = selected_boq_history_df[selected_boq_history_df["project_id"] == project_id].copy() if not selected_boq_history_df.empty else pd.DataFrame()

                top_3_share = 0.0
                if not project_boq_hist_df.empty:
                    boq_total = float(project_boq_hist_df["actual_cost"].sum())
                    top_3_share = float(project_boq_hist_df.nlargest(min(3, len(project_boq_hist_df)), "actual_cost")["actual_cost"].sum() / boq_total * 100) if boq_total else 0.0

                reasons = []
                if str(project_row["allocation_status"]) != "Balanced":
                    reasons.append("Allocation not balanced")
                if float(project_row["boq_total_cost"]) > float(project_row["total_budget"]):
                    reasons.append("BOQ exceeds budget")
                if top_3_share >= 50:
                    reasons.append("High cost concentration")
                if not project_alloc_df.empty:
                    profit_row = project_alloc_df[project_alloc_df["category"] == "Company Profit"]
                    contingency_row = project_alloc_df[project_alloc_df["category"] == "Contingency"]
                    if not profit_row.empty and float(profit_row.iloc[0]["allocation_pct"]) < 8:
                        reasons.append("Low profit margin")
                    if not contingency_row.empty and float(contingency_row.iloc[0]["allocation_pct"]) < 5:
                        reasons.append("Low contingency")

                risk_score = 0
                risk_score += 2 if str(project_row["allocation_status"]) != "Balanced" else 0
                risk_score += 2 if float(project_row["boq_total_cost"]) > float(project_row["total_budget"]) else 0
                risk_score += 2 if top_3_share >= 50 else 1 if top_3_share >= 40 else 0
                if not project_alloc_df.empty:
                    profit_row = project_alloc_df[project_alloc_df["category"] == "Company Profit"]
                    contingency_row = project_alloc_df[project_alloc_df["category"] == "Contingency"]
                    risk_score += 2 if (not profit_row.empty and float(profit_row.iloc[0]["allocation_pct"]) < 8) else 0
                    risk_score += 2 if (not contingency_row.empty and float(contingency_row.iloc[0]["allocation_pct"]) < 5) else 0

                rating = "Low Risk" if risk_score <= 2 else "Moderate Risk" if risk_score <= 5 else "High Risk"
                risk_summary_rows.append(
                    {
                        "Project ID": project_id,
                        "Project": project_row["project_name"],
                        "Risk Score": float(risk_score),
                        "Top 3 Cost Share %": round(float(top_3_share), 2),
                        "Risk Rating": rating,
                        "Key Reasons": ", ".join(reasons) if reasons else "Stable comparison profile",
                    }
                )

            risk_summary_df = pd.DataFrame(risk_summary_rows)
            if not risk_summary_df.empty:
                st.markdown("### Risk Summary for Selected Projects")
                st.dataframe(risk_summary_df, use_container_width=True, hide_index=True)

            comparison_excel_file = create_comparison_excel_report(
                selected_projects_df=selected_projects_df,
                allocation_history_df=allocation_history_df,
                comparison_history_df=comparison_history_df,
                boq_history_df=selected_boq_history_df,
                risk_summary_df=risk_summary_df,
                allocation_pivot_df=allocation_pivot_df,
                boq_pivot_df=boq_pivot_df,
                currency_symbol=currency_symbol,
            )
            st.download_button(
                label="📥 Download Historical Comparison Excel",
                data=comparison_excel_file,
                file_name="Project_Historical_Comparison.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("Select one or more saved projects to compare.")

all_saved_project_ids = saved_projects_df["id"].tolist() if not saved_projects_df.empty else []
allocation_export_history_df = load_project_allocations(all_saved_project_ids) if all_saved_project_ids else pd.DataFrame()
comparison_export_history_df = load_project_comparisons(all_saved_project_ids) if all_saved_project_ids else pd.DataFrame()
historical_boq_categories_df = load_project_boq_categories(all_saved_project_ids) if all_saved_project_ids else pd.DataFrame()
historical_overview_df = build_historical_overview(saved_projects_df)
historical_allocation_benchmark_df = build_historical_allocation_benchmark(df, allocation_export_history_df)
historical_boq_overview_df, historical_boq_benchmark_df = build_historical_boq_benchmark(
    boq_summary,
    boq_category_summary,
    historical_boq_categories_df,
)
historical_insights = generate_historical_cost_insights(
    total_budget=total_budget,
    current_budget_df=df,
    boq_df=boq_df,
    boq_summary=boq_summary,
    saved_projects_df=saved_projects_df,
    allocation_benchmark_df=historical_allocation_benchmark_df,
    boq_benchmark_df=historical_boq_benchmark_df,
)

risk_result = calculate_project_risk_score(
    total_pct=total_pct,
    allocations=allocations,
    total_budget=total_budget,
    boq_summary=boq_summary,
    boq_df=boq_df,
    allocation_benchmark_df=historical_allocation_benchmark_df,
    boq_benchmark_df=historical_boq_benchmark_df,
)
engineering_risk_insights = generate_engineering_risk_insights(
    risk_result=risk_result,
    total_budget=total_budget,
    boq_summary=boq_summary,
    saved_projects_df=saved_projects_df,
    allocation_benchmark_df=historical_allocation_benchmark_df,
    boq_benchmark_df=historical_boq_benchmark_df,
)
execution_risk_note = generate_execution_risk_note(
    project_name=project_name,
    risk_result=risk_result,
    total_budget=total_budget,
    boq_summary=boq_summary,
    saved_projects_df=saved_projects_df,
)
report_sections["Execution Risk Note"] = execution_risk_note

st.markdown("## 🧠 V16 - Historical Cost Intelligence")
if saved_projects_df.empty:
    st.info("Save project snapshots to unlock historical cost intelligence benchmarks.")
else:
    hist_metric1, hist_metric2, hist_metric3, hist_metric4 = st.columns(4)
    avg_budget = float(saved_projects_df["total_budget"].mean()) if not saved_projects_df.empty else 0.0
    avg_boq_total = float(saved_projects_df["boq_total_cost"].mean()) if not saved_projects_df.empty else 0.0
    avg_variance = float((saved_projects_df["total_budget"] - saved_projects_df["boq_total_cost"]).mean()) if not saved_projects_df.empty else 0.0
    hist_metric1.metric("Historical Avg Budget", format_currency(avg_budget, currency_symbol))
    hist_metric2.metric("Historical Avg BOQ", format_currency(avg_boq_total, currency_symbol))
    hist_metric3.metric("Current vs Hist Budget", format_currency(total_budget - avg_budget, currency_symbol))
    hist_metric4.metric("Avg Budget Variance", format_currency(avg_variance, currency_symbol))

    for insight in historical_insights:
        render_alert_box(insight)

    hist_tab1, hist_tab2, hist_tab3, hist_tab4 = st.tabs([
        "Historical Overview",
        "Allocation vs History",
        "BOQ vs History",
        "Variance & Concentration",
    ])

    with hist_tab1:
        st.dataframe(historical_overview_df, use_container_width=True, hide_index=True)
        hist_chart_df = saved_projects_df.copy()
        if not hist_chart_df.empty:
            fig_hist_3, ax_hist_3 = plt.subplots(figsize=(8, 5))
            ax_hist_3.bar(hist_chart_df["project_name"], hist_chart_df["boq_total_cost"])
            ax_hist_3.axhline(avg_boq_total, linestyle="--")
            ax_hist_3.set_title("Historical BOQ Totals")
            ax_hist_3.set_xlabel("Project")
            ax_hist_3.set_ylabel(f"Amount ({currency_code})")
            plt.xticks(rotation=45, ha="right")
            st.pyplot(fig_hist_3)
            plt.close(fig_hist_3)

    with hist_tab2:
        if historical_allocation_benchmark_df.empty:
            st.info("Historical allocation benchmarks will appear after more saved project snapshots are available.")
        else:
            display_alloc_benchmark = historical_allocation_benchmark_df.copy()
            st.dataframe(display_alloc_benchmark, use_container_width=True, hide_index=True)
            fig_hist_alloc, ax_hist_alloc = plt.subplots(figsize=(9, 5))
            benchmark_plot_df = display_alloc_benchmark.set_index("Category")[["current_allocation_pct", "historical_avg_allocation_pct"]]
            benchmark_plot_df.plot(kind="bar", ax=ax_hist_alloc)
            ax_hist_alloc.set_title("Current Allocation % vs Historical Average %")
            ax_hist_alloc.set_xlabel("Category")
            ax_hist_alloc.set_ylabel("Allocation %")
            plt.xticks(rotation=45, ha="right")
            st.pyplot(fig_hist_alloc)
            plt.close(fig_hist_alloc)

    with hist_tab3:
        if historical_boq_overview_df.empty and historical_boq_benchmark_df.empty:
            st.info("Upload a BOQ and keep saving projects to build historical BOQ category intelligence.")
        else:
            if not historical_boq_overview_df.empty:
                st.dataframe(historical_boq_overview_df, use_container_width=True, hide_index=True)
            if not historical_boq_benchmark_df.empty:
                st.dataframe(historical_boq_benchmark_df, use_container_width=True, hide_index=True)
                fig_hist_boq, ax_hist_boq = plt.subplots(figsize=(9, 5))
                boq_plot_df = historical_boq_benchmark_df.set_index("Category")[["current_actual_pct", "historical_avg_actual_pct"]]
                boq_plot_df.plot(kind="bar", ax=ax_hist_boq)
                ax_hist_boq.set_title("Current BOQ Category % vs Historical Average %")
                ax_hist_boq.set_xlabel("Category")
                ax_hist_boq.set_ylabel("Actual BOQ %")
                plt.xticks(rotation=45, ha="right")
                st.pyplot(fig_hist_boq)
                plt.close(fig_hist_boq)

    with hist_tab4:
        variance_rows = []
        current_boq_total = float(boq_summary.get("total_cost", 0.0))
        variance_rows.append(
            {
                "Metric": "Current Budget Variance",
                "Value": total_budget - current_boq_total if current_boq_total else 0.0,
            }
        )
        variance_rows.append(
            {
                "Metric": "Historical Avg Budget Variance",
                "Value": avg_variance,
            }
        )
        if not boq_df.empty and current_boq_total:
            top_item_share = float(boq_df["Total Cost"].max() / current_boq_total * 100)
            top_three_share = float(boq_df.nlargest(min(3, len(boq_df)), "Total Cost")["Total Cost"].sum() / current_boq_total * 100)
            variance_rows.append({"Metric": "Top Item Share %", "Value": top_item_share})
            variance_rows.append({"Metric": "Top 3 Items Share %", "Value": top_three_share})
        variance_df = pd.DataFrame(variance_rows)
        st.dataframe(variance_df, use_container_width=True, hide_index=True)


st.markdown("## 🚨 V17 - Engineering Risk Intelligence")
risk_metric1, risk_metric2, risk_metric3, risk_metric4 = st.columns(4)
risk_metric1.metric("Risk Score", f"{risk_result['score']}/10")
risk_metric2.metric("Risk Level", risk_result["rating"])
risk_metric3.metric("Top 3 Cost Share", f"{risk_result['concentration']['top_3_share']:.1f}%")
risk_metric4.metric("Historical Deviations", str(risk_result["deviation_count"]))

for insight in engineering_risk_insights:
    render_alert_box(insight)

risk_tab1, risk_tab2, risk_tab3 = st.tabs([
    "Risk Overview",
    "Deviation Tracker",
    "Execution Note",
])

with risk_tab1:
    risk_summary_rows = [
        {"Metric": "Risk Score", "Value": risk_result["score"]},
        {"Metric": "Risk Rating", "Value": risk_result["rating"]},
        {"Metric": "Top Item Share %", "Value": risk_result["concentration"]["top_item_share"]},
        {"Metric": "Top 3 Share %", "Value": risk_result["concentration"]["top_3_share"]},
        {"Metric": "Top 5 Share %", "Value": risk_result["concentration"]["top_5_share"]},
        {"Metric": "Concentration Risk", "Value": risk_result["concentration"]["risk_level"]},
        {"Metric": "Key Cost Drivers", "Value": risk_result["concentration"]["top_items_label"]},
    ]
    st.dataframe(pd.DataFrame(risk_summary_rows), use_container_width=True, hide_index=True)

    if risk_result["reasons"]:
        st.markdown("### Key Risk Drivers")
        for reason in risk_result["reasons"]:
            st.warning(reason)
    else:
        st.success("No major engineering risk flags were triggered by the current rules.")

with risk_tab2:
    deviation_df = risk_result["deviation_df"]
    if deviation_df.empty:
        st.success("No major historical category deviations detected.")
    else:
        st.dataframe(deviation_df, use_container_width=True, hide_index=True)
        deviation_plot_df = deviation_df.copy()
        deviation_plot_df["Label"] = deviation_plot_df["Area"] + " - " + deviation_plot_df["Category"]
        fig_dev, ax_dev = plt.subplots(figsize=(9, 5))
        ax_dev.bar(deviation_plot_df["Label"], deviation_plot_df["Variance % Points"])
        ax_dev.set_title("Historical Deviation by Cost Area")
        ax_dev.set_xlabel("Area / Category")
        ax_dev.set_ylabel("Variance % Points")
        plt.xticks(rotation=45, ha="right")
        st.pyplot(fig_dev)
        plt.close(fig_dev)

with risk_tab3:
    st.text_area("Execution Risk Note", value=execution_risk_note, height=180)
    if risk_result["concentration"]["top_items_label"] != "No BOQ uploaded":
        st.caption(f"Top BOQ cost drivers considered in the risk model: {risk_result['concentration']['top_items_label']}")

combined_export_insights = insights + historical_insights + engineering_risk_insights


excel_file = create_excel_report(
    project_name=project_name,
    project_type=project_type,
    company_name=company_name,
    client_name=client_name,
    project_reference=project_reference,
    currency_code=currency_code,
    currency_symbol=currency_symbol,
    author_name=author_name,
    today=today,
    total_budget=total_budget,
    allocations=allocations,
    df=df,
    insights=combined_export_insights,
    report_sections=report_sections,
    boq_df=boq_df,
    boq_compare_df=boq_compare_df,
    boq_summary=boq_summary,
    logo_path=logo_path_input.strip() or None,
    project_history_df=saved_projects_df,
    allocation_history_df=allocation_export_history_df,
    comparison_history_df=comparison_export_history_df,
)

pdf_file = create_pdf_report(
    project_name=project_name,
    project_type=project_type,
    company_name=company_name,
    client_name=client_name,
    project_reference=project_reference,
    currency_code=currency_code,
    currency_symbol=currency_symbol,
    author_name=author_name,
    today=today,
    total_budget=total_budget,
    allocations=allocations,
    df=df,
    insights=combined_export_insights,
    report_sections=report_sections,
    boq_df=boq_df,
    boq_compare_df=boq_compare_df,
    boq_summary=boq_summary,
)

st.markdown("## Export & Metadata")
meta_col1, meta_col2 = st.columns([1.1, 0.9])
with meta_col1:
    st.write(f"**Project Name:** {project_name}")
    st.write(f"**Project Type:** {project_type}")
    st.write(f"**Project Reference:** {project_reference}")
    st.write(f"**Client Name:** {client_name}")
    st.write(f"**Company Name:** {company_name}")
with meta_col2:
    st.write(f"**Currency:** {currency_code}")
    st.write(f"**Author:** {author_name}")
    st.write(f"**Date:** {today}")
    st.write(f"**Logo Path:** {logo_path_input if logo_path_input else 'Not provided'}")

export_col1, export_col2 = st.columns(2)
with export_col1:
    st.download_button(
        label="📥 Download Excel Report",
        data=excel_file,
        file_name="Project_Budget_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with export_col2:
    st.download_button(
        label="📄 Download PDF Report",
        data=pdf_file,
        file_name="Project_Budget_Report.pdf",
        mime="application/pdf",
        use_container_width=True,
    )

st.markdown("---")
st.markdown(
    """
<div style='text-align:center; font-size:14px; opacity:0.85; margin-top:30px;'>
<b>Project Cost Intelligence System</b><br>
Created by <b>Abdulla Zahin</b><br>
<a href='https://www.linkedin.com/in/abdulla-zahin-b4643315a/' target='_blank'>LinkedIn</a> |
<a href='https://github.com/abdulla-zahin' target='_blank'>GitHub</a>
</div>
""",
    unsafe_allow_html=True,
)
